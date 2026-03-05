"""
중국 왕홍(网红) 마케팅 자동 콘텐츠 생성기 v6
================================================================================
변경 요약 (1~8번 체크리스트)
--------------------------------------------------------------------------------
[1] 안정성 — Structured Outputs (json_schema) 적용
    - build_brief / build_creative_pack / generate_wanghong_proposal 모두
      response_format={"type":"json_schema", "json_schema": {...}} 로 변경
    - 스키마 위반 시 자동 retry 2회 → fallback json_object → 그래도 실패 시 에러

[2] 비용/속도 — st.cache_data 캐싱 + 빠른/고퀄 모드
    - build_brief_cached / build_creative_pack_cached / generate_images_cached
      (입력 해시 기반 st.cache_data 래퍼)
    - 사이드바: "⚡ 빠른 모드(이미지 2장)" / "💎 고퀄 모드(이미지 4장)" 토글

[3] 원본 보존 모드 실사용 — rembg graceful fallback + 프롬프트 강화
    - rembg 설치 시 자동 활성화 (OSError/RuntimeError 포함 광범위 예외 처리)
    - rembg 미설치 시 preserve 프롬프트에 "keep exact product shape/color/texture"
      강화 지시어 추가 → 실제 결과 차이 발생

[4] 이미지 품질 — PIL 텍스트 오버레이 (on/off 토글)
    - 이미지 생성 후 PIL로 제품명/소구점2/가격/프로모션/CTA 중국어 오버레이
    - CJK 폰트 자동 탐색 (NotoSansCJK, WenQuanYi, 맑은고딕 등)
    - 사이드바: "🖊️ 텍스트 오버레이" 토글

[5] 컴플라이언스 — 금지표현 자동 검수 + 위반 시 재생성/뱃지
    - hook/body/cta/subtitles/storyboard 전체 검수
    - 위반 시 자동 1회 재생성, 그래도 위반이면 "⚠️ 수정 필요" 뱃지 표시

[6] UX — 복사 JS 버튼 + 3안(감성/하드셀/리뷰) + 페르소나 선택
    - 제목/본문/해시태그 옆 "📋 복사" 버튼 (st.components.v1.html JS 클립보드)
    - 사이드바: 콘텐츠 톤 선택(감성/하드셀/리뷰톤)
    - 사이드바: 타겟 페르소나 선택(학생/직장인/육아/반려동물/운동/뷰티)

[7] 플랫폼 정합성 — 샤오홍슈 단일 플랫폼으로 확정
    - multiselect 제거 → 샤오홍슈 고정
    - 해시태그 수 / 캡션 길이 / 씬 구성 규칙 PLATFORMS 설정값 기반 동적 적용
    - 코드 전체에서 멀티플랫폼 가정 코드와 모순 제거

[8] 왕홍 협업 제안서 고도화
    - 3종 템플릿: DM 짧은버전(100~150자) / DM 긴버전(200~300자) / 이메일(400~700자)
    - 협업 조건 UI 입력 (콘텐츠 수/형식/납기/혜택) → 프롬프트에 반영
    - 각각 한국어/중국어 쌍 출력 + 복사 버튼
================================================================================
설치:
    pip install streamlit openai pillow python-dotenv pydantic openpyxl
    pip install rembg  # 선택사항 (원본 보존 모드 품질 향상)

.env:
    OPENAI_API_KEY=sk-...
    TEXT_MODEL=gpt-4o
    IMAGE_MODEL=gpt-image-1
"""

import os, json, base64, traceback, hashlib, re
from io import BytesIO
from typing import List, Optional, Dict, Any, Tuple

# ── [1] openpyxl ──
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── [3] rembg graceful fallback (모든 예외/충돌 포함) ──
REMBG_AVAILABLE = False
rembg_remove = None
try:
    import importlib as _il
    _rembg = _il.import_module("rembg")
    rembg_remove = _rembg.remove
    REMBG_AVAILABLE = True
except BaseException:   # ImportError/OSError/RuntimeError/SystemExit 전부 차단
    pass

import streamlit as st
import streamlit.components.v1 as components
from dotenv import load_dotenv
from PIL import Image, ImageDraw, ImageFont
from pydantic import BaseModel, Field
from openai import OpenAI

# =========================
# 0. 기본 설정
# =========================
load_dotenv()
client      = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
TEXT_MODEL  = os.getenv("TEXT_MODEL",  "gpt-4o")
IMAGE_MODEL = os.getenv("IMAGE_MODEL", "gpt-image-1")

# =========================
# 1. 플랫폼 규칙 — [7] 샤오홍슈 단일 확정
# =========================
PLATFORM = "xiaohongshu"
PLATFORM_INFO: Dict[str, Any] = {
    "name_cn":           "小红书",
    "name_kr":           "샤오홍슈",
    "caption_max_chars": 1000,
    "hashtag_count":     10,       # [7] 고정값 명시
    "scene_count":       "5~8",    # [7] 씬 수 명시
    "style":             "长文种草 / 真实体验 / 步骤清晰 / 可收藏 / 第一人称 / emoji多用",
    "image_ratio":       "3:4 세로형",
    "video_duration":    "60~90초",
}

IMAGE_SIZE_OPTIONS = {
    "정사각형 1024×1024": "1024x1024",
    "세로형 1024×1536":   "1024x1536",
    "가로형 1536×1024":   "1536x1024",
}

# [5] 카테고리별 추가 금지표현
CATEGORY_BANNED: Dict[str, List[str]] = {
    "건강식품": ["치료", "완치", "의학적 효능", "임상 입증", "FDA 승인"],
    "화장품":   ["주름 제거", "미백 보장", "성형 효과", "피부과 인증"],
    "식품":     ["다이어트 보장", "칼로리 제로 보장", "의사 추천"],
}
BASE_BANNED = ["100%", "最", "立即见效", "治愈", "无副作用", "最好", "第一"]

# [5] 금지어 한국어 설명 매핑 (비전문가 친화적 출력용)
BANNED_EXPLANATION = {
    "最":       "最 (최고·가장 등 최상급 표현 — 중국 광고법 금지)",
    "100%":     "100% (절대적 수치 보장 — 허위 광고 위험)",
    "立即见效": "立即见效 (즉각적인 효과 보장 — 과장 광고 금지)",
    "治愈":     "治愈 (치료·완치 등 의료적 주장 — 식품/가구에 사용 불가)",
    "无副作用": "无副作用 (부작용 없음 보장 — 의료 효능 주장으로 간주)",
    "最好":     "最好 (가장 좋음 — 근거 없는 최상급 금지)",
    "第一":     "第一 (1위·제일 — 증빙 없는 순위 표현 금지)",
}



# =========================
# 2. 데이터 스키마
# =========================
class Brief(BaseModel):
    brand_name:     str           = Field(...)
    product_name:   str           = Field(...)
    category:       str           = Field(...)
    target:         str           = Field(...)
    price:          Optional[str] = Field(None)
    promo:          Optional[str] = Field(None)
    key_benefits:   List[str]     = Field(default_factory=list)
    proof:          List[str]     = Field(default_factory=list)
    tone:           str           = Field("真实测评")
    banned_claims:  List[str]     = Field(default_factory=list)
    landing_action: str           = Field("点进主页领券")

class CreativePack(BaseModel):
    platform:          str
    title_cn:          str
    hook_cn:           str
    body_cn:           str
    cta_cn:            str
    hashtags:          List[str]
    thumbnail_text_cn: str
    title_kr:          str
    hook_kr:           str
    body_kr:           str
    cta_kr:            str
    thumbnail_text_kr: str
    subtitles_cn:      List[str]
    subtitles_kr:      List[str]
    storyboard:        List[Dict[str, str]]
    image_prompt:      str

# =========================
# 3. 타입 보정 유틸
# =========================
def ensure_str(v) -> str:
    if v is None: return ""
    if isinstance(v, str): return v
    if isinstance(v, list):
        return "\n".join(
            " | ".join(f"{k}:{val}" for k, val in item.items())
            if isinstance(item, dict) else str(item)
            for item in v
        )
    if isinstance(v, dict):
        return " | ".join(f"{k}:{val}" for k, val in v.items())
    return str(v)

def ensure_list(v) -> List[str]:
    if v is None: return []
    if isinstance(v, list): return [str(x).strip() for x in v if str(x).strip()]
    if isinstance(v, str):
        return [x.strip() for x in v.replace(";", ",").replace("\n", ",").split(",") if x.strip()]
    return [str(v).strip()]

def ensure_storyboard(v) -> List[Dict[str, str]]:
    if v is None: return []
    if isinstance(v, list):
        return [
            {k: str(val) for k, val in item.items()} if isinstance(item, dict)
            else {"scene": str(item), "duration": "", "visual": "", "caption_cn": "", "caption_kr": ""}
            for item in v
        ]
    return []

# =========================
# 4. OpenAI 호출 유틸 — [1] Structured Outputs with retry
# =========================
def _call_structured(system: str, user: str, schema: dict, max_retries: int = 2) -> Dict[str, Any]:
    """[1] json_schema Structured Outputs → retry → fallback json_object"""
    for attempt in range(max_retries + 1):
        try:
            resp = client.chat.completions.create(
                model=TEXT_MODEL,
                response_format={
                    "type": "json_schema",
                    "json_schema": {
                        "name": schema.get("title", "output"),
                        "strict": True,
                        "schema": schema,
                    },
                },
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user",   "content": user},
                ],
                temperature=0.85,
            )
            return json.loads(resp.choices[0].message.content)
        except Exception as e:
            if attempt == max_retries:
                # fallback to json_object
                try:
                    resp = client.chat.completions.create(
                        model=TEXT_MODEL,
                        response_format={"type": "json_object"},
                        messages=[
                            {"role": "system", "content": system + "\n\nReturn ONLY valid JSON."},
                            {"role": "user",   "content": "json\n" + user},
                        ],
                        temperature=0.85,
                    )
                    st.warning("⚠️ 구조화 출력 실패 → fallback json_object 사용 중")
                    return json.loads(resp.choices[0].message.content)
                except Exception as e2:
                    raise RuntimeError(f"Structured + fallback 모두 실패: {e2}") from e2

def pil_to_b64(img: Image.Image, fmt: str = "PNG") -> str:
    buf = BytesIO()
    img.save(buf, format=fmt)
    return base64.b64encode(buf.getvalue()).decode()

def b64_to_pil(b64: str) -> Image.Image:
    return Image.open(BytesIO(base64.b64decode(b64)))

# =========================
# 5. LLM 브리프 생성 — [1] Structured Outputs
# =========================

# [1] Brief 부분 필드 스키마 (AI가 채울 4개만)
BRIEF_PARTIAL_SCHEMA = {
    "title": "brief_partial",
    "type": "object",
    "properties": {
        "target":         {"type": "string"},
        "key_benefits":   {"type": "array", "items": {"type": "string"}},
        "tone":           {"type": "string"},
        "banned_claims":  {"type": "array", "items": {"type": "string"}},
        "landing_action": {"type": "string"},
    },
    "required": ["target", "key_benefits", "tone", "banned_claims", "landing_action"],
    "additionalProperties": False,
}

def build_brief(
    brand_name: str,
    product_name: str,
    category: str,
    price: str,
    promo: str,
    extra_notes: str,
    tone_style: str = "",
    persona: str = "",
) -> Brief:
    """[1][6] 사용자 입력값 고정 + AI는 4개 필드만 생성 + Structured Outputs"""
    cat_banned = CATEGORY_BANNED.get(category.split("/")[0] if category else "", [])
    all_banned = list(set(BASE_BANNED + cat_banned))

    sys = f"""너는 중국 B2C 왕홍 마케팅 전문 기획자다.
아래 확정 정보를 바탕으로 마케팅 브리프 JSON을 생성한다.
- 브랜드명·제품명·카테고리·가격·프로모션은 절대 변경하지 말 것.
- 오직 target, key_benefits, tone, banned_claims, landing_action 5개 키만 생성.
- key_benefits: 3~5개 배열
- banned_claims에 반드시 포함: {all_banned}
- tone 스타일 지시: {tone_style or '真实测评'}
- 타겟 페르소나 보정: {persona or '없음 (자동 추론)'}
"""
    user_prompt = (
        f"브랜드: {brand_name}\n제품: {product_name}\n카테고리: {category}\n"
        f"가격: {price}\n프로모션: {promo}\n추가 메모: {extra_notes}"
    )
    data = _call_structured(sys, user_prompt, BRIEF_PARTIAL_SCHEMA)
    data["key_benefits"]  = ensure_list(data.get("key_benefits"))
    data["banned_claims"] = list(set(ensure_list(data.get("banned_claims")) + all_banned))

    return Brief(
        brand_name=brand_name or "미입력",
        product_name=product_name,
        category=category or "미입력",
        price=price or None,
        promo=promo or None,
        target=data.get("target", ""),
        key_benefits=data["key_benefits"],
        proof=[],
        tone=data.get("tone", tone_style or "真实测评"),
        banned_claims=data["banned_claims"],
        landing_action=data.get("landing_action", "点进主页领券"),
    )

# [2] 캐싱 래퍼
@st.cache_data(show_spinner=False)
def build_brief_cached(brand_name, product_name, category, price, promo,
                       extra_notes, tone_style, persona) -> Brief:
    return build_brief(brand_name, product_name, category, price, promo,
                       extra_notes, tone_style, persona)

# =========================
# 6. 플랫폼별 마케팅 문구 — [1][5][7] Structured Outputs + 컴플라이언스
# =========================

# [1] CreativePack 스키마
CREATIVE_PACK_SCHEMA = {
    "title": "creative_pack",
    "type": "object",
    "properties": {
        "platform":          {"type": "string"},
        "title_cn":          {"type": "string"},
        "hook_cn":           {"type": "string"},
        "body_cn":           {"type": "string"},
        "cta_cn":            {"type": "string"},
        "hashtags":          {"type": "array", "items": {"type": "string"}},
        "thumbnail_text_cn": {"type": "string"},
        "title_kr":          {"type": "string"},
        "hook_kr":           {"type": "string"},
        "body_kr":           {"type": "string"},
        "cta_kr":            {"type": "string"},
        "thumbnail_text_kr": {"type": "string"},
        "subtitles_cn":      {"type": "array", "items": {"type": "string"}},
        "subtitles_kr":      {"type": "array", "items": {"type": "string"}},
        "storyboard": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "scene":      {"type": "string"},
                    "duration":   {"type": "string"},
                    "visual":     {"type": "string"},
                    "caption_cn": {"type": "string"},
                    "caption_kr": {"type": "string"},
                },
                "required": ["scene","duration","visual","caption_cn","caption_kr"],
                "additionalProperties": False,
            },
        },
        "image_prompt": {"type": "string"},
    },
    "required": [
        "platform","title_cn","hook_cn","body_cn","cta_cn","hashtags",
        "thumbnail_text_cn","title_kr","hook_kr","body_kr","cta_kr",
        "thumbnail_text_kr","subtitles_cn","subtitles_kr","storyboard","image_prompt"
    ],
    "additionalProperties": False,
}

def _check_compliance(pack: CreativePack, banned: List[str]) -> List[str]:
    """[5] 금지 표현 검수 — 위반 단어 목록 반환"""
    texts = [
        pack.title_cn, pack.hook_cn, pack.body_cn, pack.cta_cn,
        pack.title_kr, pack.hook_kr, pack.body_kr, pack.cta_kr,
    ] + pack.subtitles_cn + pack.subtitles_kr
    for scene in pack.storyboard:
        texts.extend(scene.values())
    found = []
    for b in banned:
        for t in texts:
            if b.lower() in t.lower() and b not in found:
                found.append(b)
    return found

def build_creative_pack(brief: Brief, tone_style: str = "") -> CreativePack:
    """[1][5][7] Structured Outputs + 컴플라이언스 + 단일 플랫폼"""
    p = PLATFORM_INFO
    sys = f"""너는 중국 왕홍 마케팅 콘텐츠 작가 겸 숏폼 영상 기획자다.
플랫폼: {p['name_cn']} ({p['name_kr']})
스타일: {p['style']}
콘텐츠 톤: {tone_style or brief.tone}
권장 영상 길이: {p['video_duration']}

규칙:
1. _cn 필드 → 자연스러운 중국어 마케팅 문구
2. _kr 필드 → 정확한 한국어 번역
3. body_cn: {p['caption_max_chars']}자 이내
4. hashtags: 정확히 {p['hashtag_count']}개, 각각 # 포함
5. subtitles_cn: 숏폼 자막용 중국어 (8줄 이내, 한 줄 최대 18자)
6. subtitles_kr: subtitles_cn의 한국어 번역 (같은 줄 수)
7. storyboard: {p['scene_count']}개 씬, 플랫폼 특성에 맞게
8. image_prompt: 영문 이미지 생성 프롬프트
9. 모든 단일 텍스트 필드는 반드시 문자열(str)
10. 절대 금지 표현: {brief.banned_claims}
11. 가격 {brief.price or '미입력'}, 프로모션 {brief.promo or '없음'} 반영
"""
    user = json.dumps({"brief": brief.model_dump(), "platform": PLATFORM}, ensure_ascii=False)
    data = _call_structured(sys, user, CREATIVE_PACK_SCHEMA)

    data["hashtags"]     = ensure_list(data.get("hashtags"))[:p["hashtag_count"]]
    data["subtitles_cn"] = ensure_list(data.get("subtitles_cn"))
    data["subtitles_kr"] = ensure_list(data.get("subtitles_kr"))
    data["storyboard"]   = ensure_storyboard(data.get("storyboard"))
    str_fields = ["title_cn","hook_cn","body_cn","cta_cn","thumbnail_text_cn",
                  "title_kr","hook_kr","body_kr","cta_kr","thumbnail_text_kr","image_prompt"]
    for f in str_fields:
        data[f] = ensure_str(data.get(f))
    data.setdefault("platform", PLATFORM)
    data["platform"] = PLATFORM
    pack = CreativePack(**data)

    # [5] 컴플라이언스 검수 → 위반 시 1회 재생성
    violations = _check_compliance(pack, brief.banned_claims)
    if violations:
        sys2 = sys + f"\n\n⚠️ 이전 출력에서 아래 금지 표현이 발견됨. 반드시 모두 제거하고 재생성: {violations}"
        try:
            data2 = _call_structured(sys2, user, CREATIVE_PACK_SCHEMA)
            data2["hashtags"]     = ensure_list(data2.get("hashtags"))[:p["hashtag_count"]]
            data2["subtitles_cn"] = ensure_list(data2.get("subtitles_cn"))
            data2["subtitles_kr"] = ensure_list(data2.get("subtitles_kr"))
            data2["storyboard"]   = ensure_storyboard(data2.get("storyboard"))
            for f in str_fields:
                data2[f] = ensure_str(data2.get(f))
            data2.setdefault("platform", PLATFORM)
            data2["platform"] = PLATFORM
            pack = CreativePack(**data2)
            violations = _check_compliance(pack, brief.banned_claims)
        except Exception:
            pass  # 재생성 실패 시 원본 유지

    pack._compliance_violations = violations  # type: ignore[attr-defined]
    return pack

@st.cache_data(show_spinner=False)
def build_creative_pack_cached(brief_json: str, tone_style: str) -> CreativePack:
    """[2] 캐싱 래퍼"""
    brief = Brief.model_validate_json(brief_json)
    return build_creative_pack(brief, tone_style)

# =========================
# 6-2. 한국어 → 중국어 재번역
# =========================
def retranslate_body_cn(body_kr_edited: str, brief: Brief) -> str:
    resp = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": (
                "너는 한국어→중국어 마케팅 번역 전문가다. "
                "샤오홍슈(小红书) 스타일에 맞는 자연스러운 중국어로 번역한다. "
                "직역 금지. 중국 MZ세대가 공감하는 표현을 사용한다. "
                f"금지 표현: {', '.join(brief.banned_claims)}"
            )},
            {"role": "user", "content": f"아래 한국어 마케팅 본문을 샤오홍슈 스타일 중국어로 번역해줘.\n\n{body_kr_edited}"},
        ],
        temperature=0.7, max_tokens=600,
    )
    return resp.choices[0].message.content.strip()

# =========================
# 6-3. 왕홍 협업 제안서 — [1][8] Structured Outputs + 3종 템플릿
# =========================

PROPOSAL_SCHEMA = {
    "title": "proposal",
    "type": "object",
    "properties": {
        "dm_short_kr":  {"type": "string"},
        "dm_short_cn":  {"type": "string"},
        "dm_long_kr":   {"type": "string"},
        "dm_long_cn":   {"type": "string"},
        "email_subject_kr": {"type": "string"},
        "email_subject_cn": {"type": "string"},
        "email_body_kr":    {"type": "string"},
        "email_body_cn":    {"type": "string"},
    },
    "required": [
        "dm_short_kr","dm_short_cn","dm_long_kr","dm_long_cn",
        "email_subject_kr","email_subject_cn","email_body_kr","email_body_cn"
    ],
    "additionalProperties": False,
}

def generate_wanghong_proposal(
    brief: Brief,
    collab_count: str = "1~2개",
    collab_format: str = "图文+Reel",
    collab_deadline: str = "협의 후 결정",
    collab_benefit: str = "제품 무료 제공 + 커미션",
) -> dict:
    """[8] 3종 템플릿: DM 짧은/긴 버전, 이메일 버전 (한국어+중국어)"""
    sys = """너는 한중 마케팅 협업 전문 에이전트다.
샤오홍슈 왕홍에게 보내는 3종 협업 제안서를 동시에 생성한다.
- dm_short: 100~150자 짧은 DM
- dm_long: 200~300자 긴 DM
- email: 400~700자 이메일 (제목 별도)
각각 한국어/중국어 쌍 생성. 어투: 정중하고 매력적인 비즈니스 문체."""
    user = (
        f"브랜드: {brief.brand_name} / 제품: {brief.product_name} ({brief.category})\n"
        f"타겟: {brief.target} / 핵심 강점: {', '.join(brief.key_benefits[:3])}\n"
        f"가격: {brief.price or '미입력'} / 프로모션: {brief.promo or '없음'}\n"
        f"협업 조건: 콘텐츠 {collab_count}개 / 형식: {collab_format} / 납기: {collab_deadline} / 혜택: {collab_benefit}\n"
        "위 정보로 3종 협업 제안서를 생성해줘."
    )
    return _call_structured(sys, user, PROPOSAL_SCHEMA)

# =========================
# 7. 이미지 처리 — [2][3][4]
# =========================
NO_TEXT = (
    "ABSOLUTELY NO TEXT, NO LETTERS, NO WORDS, NO FONTS, NO TYPOGRAPHY, "
    "NO LOGOS, NO WATERMARKS, NO LABELS anywhere in the image. "
    "Pure product photography only."
)

PRESERVE_PROMPT_BOOST = (
    "IMPORTANT: Keep the exact product shape, color, texture, and silhouette unchanged. "
    "Only replace/enhance the background environment. "
    "Product must be clearly recognizable and identical to the input image. "
    "Do not alter product form, proportions, or surface details. "
)

IMAGE_THEMES: List[Dict[str, str]] = [
    {
        "name_kr": "☀️ 자연채광 웜톤 거실 컷",
        "desc":    "따뜻한 오후 햇살과 우드 톤의 코지한 거실 무드",
        "style":   (
            "Interior lifestyle photography, warm afternoon sunlight streaming through sheer curtains, "
            "natural wood tones and warm beige palette, cozy and inviting living room setting, "
            "soft linen textures, potted plant in the background, "
            "Xiaohongshu interior influencer aesthetic, golden hour warmth. "
        ),
    },
    {
        "name_kr": "🛋️ 모던 미니멀리즘 컷",
        "desc":    "군더더기 없는 화이트/그레이 톤의 세련된 현대 공간",
        "style":   (
            "Modern minimalist interior photography, clean white and light grey tones, "
            "sleek contemporary furniture, uncluttered composition with intentional negative space, "
            "soft diffused studio lighting, architectural lines, "
            "Scandinavian-Korean fusion design aesthetic, premium lifestyle feel. "
        ),
    },
    {
        "name_kr": "🌃 무드 시네마틱 라운지 컷",
        "desc":    "은은한 간접 조명과 럭셔리 호텔 라운지 감성",
        "style":   (
            "Cinematic luxury lounge interior photography, warm indirect ambient lighting, "
            "deep moody tones with rich shadows, high-end hotel or boutique lounge atmosphere, "
            "plush textures, velvet and marble accents, "
            "sophisticated and aspirational mood, editorial interior style. "
        ),
    },
    {
        "name_kr": "🌿 식물테리어(플랜테리어) 컷",
        "desc":    "반려식물과 자연스럽게 어우러진 프레시한 그린 인테리어",
        "style":   (
            "Plantería interior photography, lush green houseplants of various sizes surrounding the furniture, "
            "fresh and natural biophilic design, soft daylight with green leafy shadows, "
            "terracotta pots and natural woven baskets as props, "
            "earthy and refreshing mood, urban jungle Xiaohongshu lifestyle aesthetic. "
        ),
    },
]

# [4] CJK 폰트 자동 탐색
def _find_cjk_font() -> Optional[str]:
    candidates = [
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/noto-cjk/NotoSansCJKsc-Regular.otf",
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        "/System/Library/Fonts/PingFang.ttc",
        "C:/Windows/Fonts/malgun.ttf",
        "C:/Windows/Fonts/simsun.ttc",
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None

CJK_FONT_PATH = _find_cjk_font()

def _overlay_text_on_image(
    img: Image.Image,
    brief: Brief,
    pack: Optional[CreativePack] = None,
) -> Image.Image:
    """[4] PIL로 중국어 마케팅 정보 오버레이"""
    img = img.copy().convert("RGBA")
    W, H = img.size
    draw = ImageDraw.Draw(img)

    # 반투명 하단 배너
    banner_h = int(H * 0.28)
    overlay = Image.new("RGBA", (W, banner_h), (0, 0, 0, 160))
    img.paste(overlay, (0, H - banner_h), overlay)

    try:
        if CJK_FONT_PATH:
            title_font  = ImageFont.truetype(CJK_FONT_PATH, int(W * 0.055))
            body_font   = ImageFont.truetype(CJK_FONT_PATH, int(W * 0.038))
            small_font  = ImageFont.truetype(CJK_FONT_PATH, int(W * 0.030))
        else:
            title_font = body_font = small_font = ImageFont.load_default()
    except Exception:
        title_font = body_font = small_font = ImageFont.load_default()

    y = H - banner_h + int(H * 0.015)
    pad = int(W * 0.04)

    # 제품명
    draw.text((pad, y), brief.product_name, font=title_font, fill=(255, 255, 255, 255))
    y += int(W * 0.065)

    # 핵심 소구점 2개
    for benefit in brief.key_benefits[:2]:
        draw.text((pad, y), f"✓ {benefit[:20]}", font=body_font, fill=(255, 220, 100, 255))
        y += int(W * 0.045)

    # 가격/프로모션
    if brief.price:
        price_text = f"💰 {brief.price}"
        if brief.promo:
            price_text += f"  🎁 {brief.promo[:15]}"
        draw.text((pad, y), price_text, font=body_font, fill=(255, 180, 180, 255))
        y += int(W * 0.045)

    # CTA
    cta = pack.cta_cn[:20] if pack and pack.cta_cn else "点进主页领券 →"
    draw.text((pad, y), cta, font=small_font, fill=(180, 230, 255, 255))

    return img.convert("RGB")


def _call_edit_api(img_rgba: Image.Image, prompt: str, size: str) -> Optional[Image.Image]:
    buf = BytesIO()
    img_rgba.save(buf, format="PNG")
    buf.seek(0)
    try:
        result = client.images.edit(
            model=IMAGE_MODEL,
            image=("product.png", buf, "image/png"),
            prompt=prompt,
            n=1,
            size=size,
        )
        b64 = result.data[0].b64_json
        if b64:
            return b64_to_pil(b64)
        url = result.data[0].url
        if url:
            import urllib.request
            with urllib.request.urlopen(url) as r:
                return Image.open(BytesIO(r.read()))
    except Exception as e:
        raise e
    return None


@st.cache_data(show_spinner=False)
def _generate_images_cached(
    img_b64: Optional[str],
    brief_json: str,
    image_size: str,
    synth_mode: str,
    fast_mode: bool,
) -> List[Tuple[str, str, Optional[str]]]:
    """[2] 캐싱: 결과를 (name_kr, desc, b64_or_None) 리스트로 반환"""
    brief = Brief.model_validate_json(brief_json)
    use_rembg = (synth_mode == "preserve") and REMBG_AVAILABLE
    themes = IMAGE_THEMES[:2] if fast_mode else IMAGE_THEMES  # [2] 빠른 모드

    results = []
    if img_b64:
        # 업로드 이미지 기반
        product_img = b64_to_pil(img_b64)
        if use_rembg:
            try:
                buf_in = BytesIO()
                product_img.save(buf_in, format="PNG")
                removed = rembg_remove(buf_in.getvalue())
                base_img = Image.open(BytesIO(removed)).convert("RGBA")
            except Exception:
                base_img = product_img.convert("RGBA")
        else:
            base_img = product_img.convert("RGBA")

        for theme in themes:
            # [3] preserve 모드: 프롬프트 강화
            preserve_boost = PRESERVE_PROMPT_BOOST if synth_mode == "preserve" else ""
            prompt = (
                f"{preserve_boost}{theme['style']}"
                f"Hero product: '{brief.product_name}' by brand '{brief.brand_name}'. "
                f"Target customer: {brief.target}. "
                f"Key benefits (show visually, not as text): {', '.join(brief.key_benefits[:2])}. "
                f"No medical claims. No before/after comparison. {NO_TEXT}"
            )
            try:
                img = _call_edit_api(base_img, prompt, image_size)
                results.append((theme["name_kr"], theme["desc"], pil_to_b64(img) if img else None))
            except Exception as e:
                st.warning(f"⚠️ [{theme['name_kr']}] 생성 실패: {e}")
                results.append((theme["name_kr"], theme["desc"], None))
    else:
        # 텍스트 프롬프트 기반
        for theme in themes:
            preserve_boost = (
                f"Strictly maintain the authentic appearance of '{brief.product_name}'. "
                if synth_mode == "preserve" else ""
            )
            prompt = (
                f"{preserve_boost}{theme['style']}"
                f"Product: '{brief.product_name}' by brand '{brief.brand_name}'. "
                f"Target customer: {brief.target}. "
                f"Xiaohongshu Chinese e-commerce style. No medical claims. {NO_TEXT}"
            )
            kwargs: Dict[str, Any] = {"model": IMAGE_MODEL, "prompt": prompt, "size": image_size, "n": 1}
            if "dall-e" in IMAGE_MODEL:
                kwargs["response_format"] = "b64_json"
            try:
                result = client.images.generate(**kwargs)
                b64 = result.data[0].b64_json
                results.append((theme["name_kr"], theme["desc"], b64 if b64 else None))
            except Exception as e:
                st.warning(f"⚠️ [{theme['name_kr']}] 생성 실패: {e}")
                results.append((theme["name_kr"], theme["desc"], None))

    return results  # [(name_kr, desc, b64_or_None)]


def decode_themed_images(raw: List[Tuple[str, str, Optional[str]]]) -> List[Tuple[str, str, Optional[Image.Image]]]:
    return [(n, d, b64_to_pil(b) if b else None) for n, d, b in raw]

# =========================
# 8. 엑셀 리포트 생성
# =========================
def generate_excel_report(
    brief: Brief,
    pack: CreativePack,
    proposals: dict,
    themed_images_meta: List[Tuple[str, str, Any]],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "왕홍 마케팅 기획안"

    RED    = "C0392B"; DKRED = "922B21"; LTRED = "FADBD8"
    WHITE  = "FFFFFF"; DKGRAY = "5D6D7E"

    def hfont(bold=True, color=WHITE, size=11):
        return Font(name="Arial", bold=bold, color=color, size=size)
    def bfont(bold=False, color="1A1A2E", size=10):
        return Font(name="Arial", bold=bold, color=color, size=size)
    def fill(h):
        return PatternFill("solid", start_color=h, fgColor=h)
    def center(wrap=True):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    def left(wrap=True):
        return Alignment(horizontal="left", vertical="center", wrap_text=wrap)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 62

    ws.merge_cells("A1:C1")
    ws["A1"] = f"🇨🇳 왕홍 마케팅 기획안 — {brief.product_name} ({brief.brand_name})"
    ws["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
    ws["A1"].fill = fill(RED)
    ws["A1"].alignment = center()
    ws["A1"].border = border
    ws.row_dimensions[1].height = 32

    def add_row(cat, sub, val, first=False):
        r = ws.max_row + 1
        ws.row_dimensions[r].height = max(16, min(120, 16 + str(val).count("\n") * 14))
        c1 = ws.cell(r, 1, cat); c2 = ws.cell(r, 2, sub); c3 = ws.cell(r, 3, str(val))
        c1.font = hfont(bold=first, color=WHITE if first else DKRED, size=11 if first else 10)
        c1.fill = fill(DKRED if first else LTRED)
        c1.alignment = center(); c1.border = border
        c2.font = hfont(bold=True, color=DKRED, size=10)
        c2.fill = fill(LTRED); c2.alignment = left(); c2.border = border
        c3.font = bfont(); c3.fill = fill(WHITE); c3.alignment = left(); c3.border = border

    CAT1 = "1. 캠페인 개요"
    add_row(CAT1, "브랜드명",   brief.brand_name,  True)
    add_row(CAT1, "제품명",     brief.product_name)
    add_row(CAT1, "카테고리",   brief.category)
    add_row(CAT1, "타겟 고객",  brief.target)
    add_row(CAT1, "핵심 소구점", "\n".join(f"• {b}" for b in brief.key_benefits))
    add_row(CAT1, "가격",       brief.price or "미입력")
    add_row(CAT1, "프로모션",   brief.promo or "미입력")
    add_row(CAT1, "마케팅 톤",  brief.tone)
    add_row(CAT1, "금지 표현",  " / ".join(brief.banned_claims))

    CAT2 = "2. 시각/이미지 전략"
    for i, (n, d, _) in enumerate(themed_images_meta or []):
        add_row(CAT2, n, d, i == 0)

    CAT3 = "3. 플랫폼 커뮤니케이션\n(샤오홍슈)"
    add_row(CAT3, "제목 (CN)",       pack.title_cn,       True)
    add_row(CAT3, "제목 (KR)",       pack.title_kr)
    add_row(CAT3, "핵심 캡션 (CN)",  pack.hook_cn)
    add_row(CAT3, "핵심 캡션 (KR)",  pack.hook_kr)
    add_row(CAT3, "썸네일 문구 (CN)",pack.thumbnail_text_cn)
    add_row(CAT3, "썸네일 문구 (KR)",pack.thumbnail_text_kr)
    add_row(CAT3, "본문 (중국어)",    pack.body_cn)
    add_row(CAT3, "본문 (한국어)",    pack.body_kr)
    add_row(CAT3, "CTA (CN)",        pack.cta_cn)
    add_row(CAT3, "CTA (KR)",        pack.cta_kr)
    add_row(CAT3, "해시태그",         " ".join(pack.hashtags))

    CAT4 = "4. 영상 & 바이럴\n실행 계획"
    add_row(CAT4, "숏폼 자막 (CN)", "\n".join(f"{i}. {s}" for i,s in enumerate(pack.subtitles_cn,1)), True)
    add_row(CAT4, "숏폼 자막 (KR)", "\n".join(f"{i}. {s}" for i,s in enumerate(pack.subtitles_kr,1)))
    for scene in pack.storyboard:
        add_row(CAT4, "스토리보드",
                f"[{scene.get('scene','')}] {scene.get('duration','')}\n"
                f"화면: {scene.get('visual','')}\nCN: {scene.get('caption_cn','')}\nKR: {scene.get('caption_kr','')}")
    add_row(CAT4, "왕홍 제안서 DM (KR)", proposals.get("dm_long_kr","미생성"))
    add_row(CAT4, "왕홍 제안서 DM (CN)", proposals.get("dm_long_cn","미생성"))

    CAT5 = "5. 기대 효과"
    add_row(CAT5, "KPI 노출 목표",
            "• 샤오홍슈 순 노출 50만 회 이상\n• 콘텐츠 저장(收藏) 7일 내 1,000건\n• 팔로워 +15% 이상", True)
    add_row(CAT5, "전환율 목표", "링크 클릭 → 구매 전환율 3.5% / ROI 목표 3배")
    add_row(CAT5, "브랜드 인지도", "키워드 검색량 2배 증가 / 호감도 70% 이상")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# =========================
# 9. [6] 복사 버튼 JS 헬퍼
# =========================
def copy_button(text: str, key: str):
    """[6] 클립보드 복사 버튼 — base64 인코딩으로 JS 삽입 문제 완전 차단.
    줄바꿈·따옴표·백틱·중국어 등 모든 특수문자에서 안전하게 동작."""
    b64 = base64.b64encode(text.encode("utf-8")).decode("ascii")
    components.html(
        f"""
        <button id="btn_{key}"
            data-b64="{b64}"
            onclick="
                var raw = this.getAttribute('data-b64');
                var bytes = Uint8Array.from(atob(raw), c => c.charCodeAt(0));
                var decoded = new TextDecoder('utf-8').decode(bytes);
                navigator.clipboard.writeText(decoded).then(function() {{
                    var btn = document.getElementById('btn_{key}');
                    btn.textContent = '✅ 복사됨!';
                    setTimeout(function() {{ btn.textContent = '📋 복사'; }}, 1500);
                }}).catch(function() {{
                    var btn = document.getElementById('btn_{key}');
                    btn.textContent = '❌ 복사 실패';
                    setTimeout(function() {{ btn.textContent = '📋 복사'; }}, 1500);
                }});
            "
            style="font-size:12px;padding:4px 12px;border:1px solid #ccc;
                   border-radius:4px;background:#fff;cursor:pointer;margin-top:2px;">
        📋 복사</button>
        """,
        height=36,
    )

# =========================
# 10. Streamlit UI
# =========================
st.set_page_config(
    page_title="🇨🇳 중국 왕홍 마케팅 자동 생성기 v6",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── session_state 초기화 ──────────────────────
_SS_DEFAULTS = {
    "magic_prompt": "",
    "retranslated": {},
    "extra_notes": "",
    "wanghong_proposals": {},
    "is_generated": False,
    "ss_brief": None,
    "ss_pack": None,
    "ss_themed_raw": [],
    "ss_use_rembg": False,
    "ss_uploaded": False,
    "ss_product_query": "",
}
for k, v in _SS_DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── 사이드바 ──────────────────────────────────
with st.sidebar:
    st.title("⚙️ 설정")

    # [7] 샤오홍슈 단일 플랫폼 고정
    st.info(f"📱 플랫폼: **{PLATFORM_INFO['name_kr']}** (고정)")

    image_size_label = st.selectbox("이미지 규격", list(IMAGE_SIZE_OPTIONS.keys()), index=0)
    image_size = IMAGE_SIZE_OPTIONS[image_size_label]

    st.markdown("---")

    # [2] 빠른/고퀄 모드
    speed_mode = st.radio(
        "생성 모드",
        ["⚡ 빠른 모드 (이미지 2장)", "💎 고퀄 모드 (이미지 4장)"],
        index=1, key="speed_mode",
    )
    fast_mode = "빠른" in speed_mode

    # [4] 텍스트 오버레이
    overlay_on = st.toggle("🖊️ 중국어 텍스트 오버레이", value=False, key="overlay_on")
    if overlay_on and not CJK_FONT_PATH:
        st.warning("⚠️ CJK 폰트 미발견 → 기본 폰트로 대체됩니다.")

    st.markdown("---")

    # [6] 타겟 페르소나 — 자유 텍스트 입력
    st.markdown("**🎯 타겟 고객**")
    persona = st.text_input(
        "타겟 고객",
        placeholder="예: 인테리어에 관심 많은 신혼부부, 좁은 방을 꾸미는 1인가구",
        key="persona",
        label_visibility="collapsed",
    )

    # [6] 마케팅 톤앤매너 — 자유 텍스트 입력
    st.markdown("**🎨 원하는 마케팅 느낌**")
    tone_style = st.text_input(
        "마케팅 톤앤매너",
        placeholder="예: 고급스럽고 세련된 느낌, 따뜻하고 포근한 감성",
        key="tone_style_input",
        label_visibility="collapsed",
    )
    st.caption("💡 마케팅 용어를 몰라도 괜찮습니다. '내돈내산 후기처럼', '감성적인 카페 분위기' 등 원하는 느낌을 편하게 적어주세요.")

    st.markdown("---")
    st.markdown("**모델 설정** (`.env`)")
    st.code(f"TEXT_MODEL  = {TEXT_MODEL}\nIMAGE_MODEL = {IMAGE_MODEL}", language="bash")
    if REMBG_AVAILABLE:
        st.success("✅ rembg 설치됨 — 원본 보존 모드 활성")
    else:
        st.info("ℹ️ rembg 미설치 — 프롬프트 강화 모드로 대체")

# ── 메인 헤더 ─────────────────────────────────
st.title("🇨🇳 중국 왕홍 마케팅 자동 생성기 v6")
st.caption(
    "① 제품 정보 입력 → ② AI 마케팅 브리프 분석 → "
    "③ 샤오홍슈 마케팅 이미지 4테마 생성 → ④ 중국어/한국어 콘텐츠 자동 생성 → ⑤ 엑셀 기획안 다운로드"
)
st.divider()

# ── 입력 영역 ─────────────────────────────────
inp_a, inp_b = st.columns([1, 1], gap="large")

with inp_a:
    st.subheader("📦 제품 기본 정보")
    col_a1, col_a2 = st.columns(2, gap="small")
    with col_a1:
        input_brand   = st.text_input("브랜드명 *", placeholder="예: 비타제주", key="input_brand")
    with col_a2:
        input_product = st.text_input("제품명 *", placeholder="예: 제주 감귤 비타민 젤리", key="input_product")
    input_category    = st.text_input("카테고리", placeholder="예: 건강식품/간식", key="input_category")
    col_a3, col_a4 = st.columns(2, gap="small")
    with col_a3:
        input_price   = st.text_input("가격 (위안화)", placeholder="예: 69위안", key="input_price")
    with col_a4:
        input_promo   = st.text_input("프로모션 내용", placeholder="예: 신제품 출시 기념 20% 할인", key="input_promo")

    product_query = input_product

    st.markdown("")
    if st.button("✨ AI 마케팅 포인트 자동 추천받기", key="magic_btn", use_container_width=True):
        if not input_product.strip():
            st.warning("제품명을 먼저 입력해주세요.")
        else:
            with st.spinner("AI가 마케팅 포인트를 분석 중입니다..."):
                try:
                    # 사이드바 입력값 반영 (비어있으면 기본값 사용)
                    _persona_hint = (
                        f"반드시 이 타겟 고객에 맞게 작성: {persona}"
                        if persona.strip() else "타겟 고객: 자동 추론"
                    )
                    _tone_hint = (
                        f"반드시 이 마케팅 느낌/톤으로 작성: {tone_style}"
                        if tone_style.strip() else "마케팅 톤: 자연스럽고 공감되는 문체"
                    )
                    resp = client.chat.completions.create(
                        model="gpt-4o-mini",
                        messages=[
                            {"role": "system", "content": (
                                "너는 중국 샤오홍슈 마케팅 전략가다. "
                                "확정된 제품 정보와 사용자가 지정한 타겟·톤을 최우선으로 반영해 "
                                "마케팅 전략 명세서를 작성한다. "
                                "브랜드명·제품명·가격·프로모션은 절대 변경하지 않는다. "
                                "반드시 한국어로 작성. 근거 없는 효능·의료 주장 금지. "
                                f"{_persona_hint}. {_tone_hint}."
                            )},
                            {"role": "user", "content": (
                                f"브랜드: {input_brand}\n제품명: {input_product}\n"
                                f"카테고리: {input_category}\n가격: {input_price}\n프로모션: {input_promo}\n"
                                f"타겟 고객: {persona if persona.strip() else '(자동 추론)'}\n"
                                f"원하는 마케팅 느낌: {tone_style if tone_style.strip() else '(자동 추론)'}\n\n"
                                "위 정보를 바탕으로 샤오홍슈 마케팅 전략 명세서를 5가지 항목으로 작성해줘 "
                                "(**볼드 제목** + 내용, 각 항목 2~3줄):\n"
                                "**1. 핵심 소구점** (이 타겟에게 가장 와닿는 강점 3가지)\n"
                                "**2. 공간/감각적 특징** (제품이 놓인 공간·질감·분위기 묘사)\n"
                                "**3. 타겟 페인포인트** (위 타겟 고객의 구체적인 고민과 이 제품이 해결하는 방식)\n"
                                "**4. 시각적 무드 & 컬러** (위 톤앤매너에 맞는 이미지 방향)\n"
                                "**5. 추천 콘텐츠 톤앤매너** (위 마케팅 느낌을 살린 샤오홍슈 문체 방향)"
                            )},
                        ],
                        temperature=0.75, max_tokens=800,
                    )
                    result_text = resp.choices[0].message.content.strip()
                    st.session_state.magic_prompt = result_text
                    st.session_state.extra_notes  = result_text
                    st.success("✅ 마케팅 포인트 자동 추천 완료! 아래 내용을 확인하고 수정하세요.")
                except Exception as e:
                    st.error(f"AI 추천 실패: {e}")

    extra_notes = st.text_area(
        "📄 AI 추천 마케팅 메모 (자동 채워짐 / 직접 수정 가능)",
        placeholder="위 버튼으로 자동 추천받거나, 타겟·톤앤매너 등을 직접 입력하세요.",
        height=150, key="extra_notes",
    )

with inp_b:
    st.subheader("🖼️ 제품 이미지 업로드")
    uploaded_file = st.file_uploader(
        "제품 사진을 업로드하면 AI가 테마 마케팅 이미지를 생성합니다.",
        type=["jpg","jpeg","png","webp"], key="product_image",
    )
    if uploaded_file:
        preview_img = Image.open(uploaded_file)
        st.image(preview_img, caption="업로드된 제품 이미지", use_container_width=True)
        st.success("✅ 이미지 업로드 완료!")
    else:
        st.info("📌 이미지 미업로드 시 텍스트 기반으로 이미지를 생성합니다.")

    st.markdown("---")
    st.markdown("**🎛️ 이미지 합성 모드 선택**")
    synth_mode = st.radio(
        "합성 방식",
        options=["preserve", "creative"],
        format_func=lambda x: (
            "✂️ 제품 원본 보존 모드 (누끼 추출 + 배경 합성)" if x == "preserve"
            else "🎨 AI 창의적 재생성 모드 (형태 변형 가능)"
        ),
        index=0, key="synth_mode", label_visibility="collapsed",
    )
    if synth_mode == "preserve":
        rembg_badge = "✅ rembg 활성" if REMBG_AVAILABLE else "🔧 프롬프트 강화 모드"
        st.caption(f"✂️ 원본 보존 모드 ({rembg_badge}): 제품 형태·색상 최대 유지. 배경만 교체합니다.")
    else:
        st.caption("🎨 창의적 재생성 모드: AI가 자유롭게 재해석합니다. 감각적인 연출 컷에 적합합니다.")

st.markdown("")
_can = bool(input_product.strip()) and bool(input_brand.strip())
submitted = st.button(
    "🚀 마케팅 콘텐츠 자동 생성",
    type="primary", disabled=not _can,
    use_container_width=True, key="gen_btn",
)
if not input_product.strip() or not input_brand.strip():
    st.caption("👆 브랜드명과 제품명을 입력하면 버튼이 활성화됩니다.")

st.divider()

# ── 생성 로직 (API 호출만) ─────────────────────
if submitted and _can:
    if not os.getenv("OPENAI_API_KEY"):
        st.error("❌ OPENAI_API_KEY가 설정되지 않았습니다.")
        st.stop()

    st.session_state.retranslated = {}
    st.session_state.wanghong_proposals = {}

    try:
        with st.status("🔍 제품 마케팅 브리프 분석 중...", expanded=False) as s:
            brief = build_brief_cached(
                input_brand, input_product, input_category,
                input_price, input_promo, extra_notes, tone_style, persona
            )
            st.session_state.ss_brief = brief
            s.update(label="✅ 브리프 분석 완료", state="complete")

        n_themes = 2 if fast_mode else 4
        with st.status(f"🎨 AI 마케팅 이미지 {n_themes}가지 테마 생성 중...", expanded=False) as s:
            img_b64 = None
            if uploaded_file:
                uploaded_file.seek(0)
                img_b64 = pil_to_b64(Image.open(uploaded_file))
            themed_raw = _generate_images_cached(
                img_b64, brief.model_dump_json(),
                image_size, synth_mode, fast_mode,
            )
            st.session_state.ss_themed_raw  = themed_raw
            st.session_state.ss_use_rembg   = REMBG_AVAILABLE and synth_mode == "preserve"
            st.session_state.ss_uploaded    = bool(uploaded_file)
            ok = sum(1 for _, _, b in themed_raw if b)
            s.update(label=f"✅ 이미지 생성 완료 ({ok}/{n_themes}장)", state="complete" if ok else "error")

        with st.status("✍️ 샤오홍슈 콘텐츠 생성 중...", expanded=False) as s:
            pack = build_creative_pack_cached(brief.model_dump_json(), tone_style)
            st.session_state.ss_pack = pack
            s.update(label="✅ 콘텐츠 생성 완료", state="complete")

        st.session_state.ss_product_query = input_product
        st.session_state.is_generated = True
        st.success("🎉 모든 마케팅 콘텐츠 생성이 완료되었습니다!")

    except Exception as e:
        st.error(f"❌ 오류 발생: {e}")
        with st.expander("상세 오류 내용 보기"):
            st.code(traceback.format_exc(), language="python")
        st.info("💡 `pip install -U openai` 업데이트 / `.env`에서 IMAGE_MODEL 확인")

# ══════════════════════════════════════════
# 결과 표시 — session_state 기반
# ══════════════════════════════════════════
if st.session_state.get("is_generated", False):
    brief         = st.session_state.ss_brief
    themed_raw    = st.session_state.ss_themed_raw
    pack          = st.session_state.ss_pack
    use_rembg     = st.session_state.ss_use_rembg
    themed_images = decode_themed_images(themed_raw)

    st.divider()

    # ── 제품 브리프 요약 ──────────────────────
    with st.expander("📋 자동 분석된 제품 브리프", expanded=False):
        b1, b2 = st.columns(2)
        with b1:
            st.markdown(f"**브랜드명**: {brief.brand_name}")
            st.markdown(f"**제품명**: {brief.product_name}")
            st.markdown(f"**카테고리**: {brief.category}")
            st.markdown(f"**타겟 고객**: {brief.target}")
            st.markdown(f"**가격**: {brief.price or '미입력'}")
            st.markdown(f"**프로모션**: {brief.promo or '미입력'}")
        with b2:
            st.markdown(f"**마케팅 톤**: {brief.tone}")
            st.markdown(f"**CTA**: {brief.landing_action}")
            for b in brief.key_benefits:
                st.markdown(f"  - {b}")
            st.code(" / ".join(brief.banned_claims))

    # ── AI 이미지 갤러리 ──────────────────────
    n_shown = len(themed_images)
    mode_label = (
        "✂️ 원본 보존 (rembg 누끼)" if use_rembg
        else ("🔧 원본 보존 (프롬프트 강화)" if st.session_state.ss_uploaded and synth_mode == "preserve"
              else ("🎨 AI 창의적 재생성" if st.session_state.ss_uploaded else "📝 텍스트 프롬프트"))
    )
    st.subheader(f"🖼️ AI 마케팅 이미지 갤러리 ({n_shown}가지 테마)")
    st.caption(f"합성 방식: **{mode_label}**  |  텍스트 오버레이: {'ON 🖊️' if overlay_on else 'OFF'}")

    for row_start in range(0, len(themed_images), 2):
        row_items = themed_images[row_start:row_start+2]
        cols = st.columns(2, gap="medium")
        for col, (theme_name, theme_desc, img) in zip(cols, row_items):
            with col:
                with st.container(border=True):
                    st.markdown(f"**{theme_name}**")
                    st.caption(theme_desc)
                    if img:
                        # [4] 텍스트 오버레이
                        display_img = _overlay_text_on_image(img, brief, pack) if overlay_on else img
                        st.image(display_img, use_container_width=True)
                        buf = BytesIO()
                        display_img.save(buf, format="PNG")
                        safe_name = re.sub(r"[^\w가-힣]", "", theme_name).strip().replace(" ", "_")
                        st.download_button(
                            f"⬇️ {theme_name} 다운로드",
                            data=buf.getvalue(),
                            file_name=f"{brief.product_name}_{safe_name}.png",
                            mime="image/png",
                            key=f"dl_{row_start}_{safe_name}",
                            use_container_width=True,
                        )
                    else:
                        st.warning("이 테마 이미지 생성에 실패했습니다.")

    st.info("💡 Canva 등에서 추가 편집하거나, 사이드바에서 이미지 규격을 변경 후 재생성하세요.")
    st.divider()

    # ── 플랫폼 마케팅 콘텐츠 ──────────────────
    # [5] 컴플라이언스 결과 — 비전문가 친화적 출력
    violations = getattr(pack, "_compliance_violations", [])
    if violations:
        explained = [BANNED_EXPLANATION.get(v, v) for v in violations]
        st.error(
            "🚨 **[중국 광고법 위반 주의]** "
            "샤오홍슈 계정 차단 위험이 있는 금지어가 발견되었습니다! "
            "아래 단어를 본문에서 지우거나 부드러운 표현으로 수정해 주세요.\n\n"
            + "\n".join(f"- {e}" for e in explained)
        )
    else:
        st.success("✅ 중국 광고법 필수 검수 통과 (안전)")

    st.subheader("🧩 샤오홍슈 마케팅 콘텐츠")
    st.markdown(f"### 📝 {PLATFORM_INFO['name_kr']} 맞춤 마케팅 문구")
    mc1, mc2 = st.columns(2, gap="large")

    current_body_cn = st.session_state.retranslated.get(PLATFORM, pack.body_cn)

    with mc1:
        st.markdown("#### 🇨🇳 중국어 번역본")
        with st.container(border=True):
            st.markdown(f"**제목** | {pack.title_cn}")
            copy_button(pack.title_cn, "cp_title_cn")  # [6]
            st.markdown(f"**핵심 캡션** | {pack.hook_cn}")
            st.markdown(f"**CTA** | {pack.cta_cn}")
            st.markdown(f"**썸네일 문구** | {pack.thumbnail_text_cn}")
        st.markdown("**본문 (중국어)** — 클릭 후 Ctrl+A 복사")
        st.text_area("중국어 본문", value=current_body_cn, height=220, label_visibility="collapsed")
        copy_button(current_body_cn, "cp_body_cn")  # [6]
        st.markdown("**해시태그**")
        st.text_area("해시태그", value=" ".join(pack.hashtags), height=68, label_visibility="collapsed")
        copy_button(" ".join(pack.hashtags), "cp_hashtags")  # [6]

    with mc2:
        st.markdown("#### 🇰🇷 한국어 원문 수정 → 중국어 재번역")
        with st.container(border=True):
            st.markdown(f"**제목** | {pack.title_kr}")
            st.markdown(f"**핵심 캡션** | {pack.hook_kr}")
            st.markdown(f"**CTA** | {pack.cta_kr}")
            st.markdown(f"**썸네일 문구** | {pack.thumbnail_text_kr}")
        st.markdown("**본문 (한국어) — 직접 수정 후 재번역**")
        if f"kr_edit_{PLATFORM}" not in st.session_state:
            st.session_state[f"kr_edit_{PLATFORM}"] = pack.body_kr
        edited_kr = st.text_area(
            "한국어 본문 수정", height=220,
            key=f"kr_edit_{PLATFORM}", label_visibility="collapsed",
        )
        if st.button("🔄 수정한 한국어로 중국어 재번역하기",
                     key=f"retrans_{PLATFORM}", use_container_width=True, type="secondary"):
            with st.spinner("중국어로 재번역 중..."):
                try:
                    new_cn = retranslate_body_cn(edited_kr, brief)
                    st.session_state.retranslated[PLATFORM] = new_cn
                    st.success("✅ 재번역 완료!")
                    st.rerun()
                except Exception as e:
                    st.error(f"재번역 실패: {e}")

    st.divider()

    # ── 자막 / 스토리보드 ─────────────────────
    st.markdown("### 🧷 숏폼 영상 자막")
    sub_c1, sub_c2 = st.columns(2)
    with sub_c1:
        st.markdown("**🇨🇳 중국어 자막**")
        subs_cn_text = "\n".join(f"{i}. {cn}" for i, cn in enumerate(pack.subtitles_cn, 1))
        st.text_area("CN자막", value=subs_cn_text, height=160, label_visibility="collapsed")
        copy_button(subs_cn_text, "cp_subs_cn")
    with sub_c2:
        st.markdown("**🇰🇷 한국어 번역**")
        subs_kr_text = "\n".join(f"{i}. {kr}" for i, kr in enumerate(pack.subtitles_kr, 1))
        st.text_area("KR자막", value=subs_kr_text, height=160, label_visibility="collapsed")

    st.divider()
    with st.expander("🎬 숏폼 영상 제작 기획안 (씬별 스토리보드)", expanded=True):
        st.caption(f"권장 길이: {PLATFORM_INFO['video_duration']}")
        if pack.storyboard:
            for i, scene in enumerate(pack.storyboard, 1):
                with st.container(border=True):
                    sc1, sc2 = st.columns([1, 2])
                    with sc1:
                        st.markdown(f"**{scene.get('scene', f'씬 {i}')}**")
                        st.markdown(f"⏱ {scene.get('duration','')}")
                    with sc2:
                        st.markdown(f"📷 **화면**: {scene.get('visual','')}")
                        st.markdown(f"🇨🇳 **자막**: {scene.get('caption_cn','')}")
                        st.markdown(f"🇰🇷 **번역**: {scene.get('caption_kr','')}")
        else:
            st.info("스토리보드 생성에 실패했습니다.")

    # ── [8] 왕홍 협업 제안서 고도화 ──────────
    st.divider()
    with st.expander("💌 왕홍 협업 제안서 (DM 3종 + 이메일)", expanded=False):
        st.caption("3가지 템플릿 중 상황에 맞는 것을 선택해 바로 복사하세요.")

        # [8] 협업 조건 UI 입력
        st.markdown("**📋 협업 조건 설정**")
        c1, c2, c3, c4 = st.columns(4)
        with c1: collab_count   = st.text_input("콘텐츠 수", value="1~2개", key="collab_count")
        with c2: collab_format  = st.text_input("형식", value="图文+Reel", key="collab_format")
        with c3: collab_deadline= st.text_input("납기", value="협의 후 결정", key="collab_deadline")
        with c4: collab_benefit = st.text_input("제공 혜택", value="제품 무료 제공 + 커미션", key="collab_benefit")

        if st.button("✉️ 3종 협업 제안서 생성하기", key="proposal_btn", use_container_width=True):
            with st.spinner("협업 제안서를 작성 중입니다..."):
                try:
                    proposals = generate_wanghong_proposal(
                        brief, collab_count, collab_format, collab_deadline, collab_benefit
                    )
                    st.session_state.wanghong_proposals = proposals
                    st.success("✅ 제안서 생성 완료!")
                except Exception as e:
                    st.error(f"제안서 생성 실패: {e}")

        proposals = st.session_state.get("wanghong_proposals", {})
        if proposals:
            prop_tab1, prop_tab2, prop_tab3 = st.tabs(
                ["📩 DM 짧은버전 (100~150자)", "📨 DM 긴버전 (200~300자)", "📧 이메일 버전"]
            )
            with prop_tab1:
                pc1, pc2 = st.columns(2)
                with pc1:
                    st.markdown("🇰🇷 한국어")
                    st.text_area("dm_short_kr", value=proposals.get("dm_short_kr",""), height=120, label_visibility="collapsed")
                    copy_button(proposals.get("dm_short_kr",""), "cp_dm_short_kr")
                with pc2:
                    st.markdown("🇨🇳 중국어")
                    st.text_area("dm_short_cn", value=proposals.get("dm_short_cn",""), height=120, label_visibility="collapsed")
                    copy_button(proposals.get("dm_short_cn",""), "cp_dm_short_cn")
            with prop_tab2:
                pc1, pc2 = st.columns(2)
                with pc1:
                    st.markdown("🇰🇷 한국어")
                    st.text_area("dm_long_kr", value=proposals.get("dm_long_kr",""), height=180, label_visibility="collapsed")
                    copy_button(proposals.get("dm_long_kr",""), "cp_dm_long_kr")
                with pc2:
                    st.markdown("🇨🇳 중국어")
                    st.text_area("dm_long_cn", value=proposals.get("dm_long_cn",""), height=180, label_visibility="collapsed")
                    copy_button(proposals.get("dm_long_cn",""), "cp_dm_long_cn")
            with prop_tab3:
                pc1, pc2 = st.columns(2)
                with pc1:
                    st.markdown("🇰🇷 한국어")
                    st.markdown(f"**제목**: {proposals.get('email_subject_kr','')}")
                    st.text_area("email_body_kr", value=proposals.get("email_body_kr",""), height=260, label_visibility="collapsed")
                    copy_button(proposals.get("email_body_kr",""), "cp_email_kr")
                with pc2:
                    st.markdown("🇨🇳 중국어")
                    st.markdown(f"**主题**: {proposals.get('email_subject_cn','')}")
                    st.text_area("email_body_cn", value=proposals.get("email_body_cn",""), height=260, label_visibility="collapsed")
                    copy_button(proposals.get("email_body_cn",""), "cp_email_cn")
        else:
            st.info("위 버튼을 눌러 협업 제안서를 생성하세요.")

    # ── 종합 기획안 다운로드 ──────────────────
    st.divider()
    st.subheader("📊 종합 광고 기획안 다운로드")
    st.caption("💡 엑셀 파일을 열어 내용을 확인·수정해 실무에 바로 활용하세요. Excel / Google Sheets / Numbers 모두 지원됩니다.")

    fin_col1, fin_col2 = st.columns(2, gap="medium")
    with fin_col1:
        proposals_for_report = st.session_state.get("wanghong_proposals", {})
        excel_bytes = generate_excel_report(brief, pack, proposals_for_report, themed_images)
        st.download_button(
            "📊 최종 종합 광고 기획안 다운로드 (Excel)",
            data=excel_bytes,
            file_name=f"왕홍마케팅_기획안_{brief.product_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary",
        )
    with fin_col2:
        export = {
            "brief":   brief.model_dump(),
            "pack":    pack.model_dump(),
            "proposals": proposals_for_report,
        }
        st.download_button(
            "⬇️ 원본 데이터 JSON 다운로드",
            data=json.dumps(export, ensure_ascii=False, indent=2).encode("utf-8"),
            file_name=f"wanghong_{brief.product_name}_output.json",
            mime="application/json", use_container_width=True,
        )