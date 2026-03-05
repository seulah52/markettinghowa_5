"""
중국 왕홍 마케팅 자동 콘텐츠 생성기 v7
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
app12.py 기반 완전 병합 + 추가 기능
- 6개 플랫폼 선택 (눈에 띄는 카드 UI, 중국 초보 친화)
- remove.bg → rembg fallback 배경제거
- GPT-4o Vision 이미지 분석 → 한국어 자동 기입
- app12.py IMAGE_THEMES + _call_edit_api 배경 로직
- DeepSeek 문구/해시태그 추천 (한국어 번역 표시)
- PIL 자동 줄바꿈 문구 합성 (한국어 합성 안 됨)
- 영상 프롬프트 (씬별 + 전체 복사) + AI 영상 사이트 링크
- 왕홍 협업 제안서 3종 (DM짧/DM긴/이메일)
- 엑셀/JSON 다운로드
- 버튼식 단계별 이동 (STEP 1~4)
- 모듈1 데이터 연동 로직
"""

import os, json, base64, traceback, re, hashlib
import requests
from io import BytesIO
from typing import List, Optional, Dict, Any, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side

# rembg graceful fallback
REMBG_AVAILABLE = False
rembg_remove = None
try:
    import importlib as _il
    _rembg = _il.import_module("rembg")
    rembg_remove = _rembg.remove
    REMBG_AVAILABLE = True
except BaseException:
    pass

import streamlit as st
import streamlit.components.v1 as components
from dotenv import load_dotenv
from PIL import Image, ImageDraw, ImageFont
from pydantic import BaseModel, Field
from openai import OpenAI

load_dotenv()

openai_client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
deepseek_client = OpenAI(
    api_key=os.getenv("DEEPSEEK_API_KEY", ""),
    base_url="https://api.deepseek.com"
)
REMOVE_BG_KEY = os.getenv("REMOVE_BG_API_KEY", "")

TEXT_MODEL  = os.getenv("TEXT_MODEL",  "gpt-4o")
IMAGE_MODEL = os.getenv("IMAGE_MODEL", "gpt-image-1")

IMAGE_SIZE_OPTIONS = {
    "정사각형 1024×1024 (SNS 피드)": "1024x1024",
    "세로형 1024×1536 (샤오홍슈/상세)": "1024x1536",
    "가로형 1536×1024 (배너)": "1536x1024",
}

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 1. 플랫폼 데이터 (6개, 중국 초보 친화)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PLATFORMS: Dict[str, Dict] = {
    "xiaohongshu": {
        "name_kr": "샤오홍슈", "name_cn": "小红书", "icon": "📕",
        "monthly_users": "월 3억+ 명",
        "main_users": "20~35세 도시 여성",
        "best_category": "뷰티 / 인테리어 / 패션 / 음식",
        "content_type": "감성 사진 + 솔직 후기 글",
        "beginner_tip": "👶 첫 진출 추천 No.1! 한국 제품 선호도가 가장 높음",
        "desc": "중국판 인스타그램+블로그. 진정성 있는 후기와 감성 화보가 핵심.",
        "color": "#e91e63",
        "style": "长文种草 / 真实体验 / 步骤清晰 / 可收藏 / 第一人称 / emoji多用",
        "hashtag_count": 10,
        "scene_count": "5~8",
        "video_duration": "60~90초",
        "caption_max": 1000,
    },
    "taobao": {
        "name_kr": "타오바오", "name_cn": "淘宝", "icon": "🛍️",
        "monthly_users": "월 8억+ 명",
        "main_users": "전 연령, 가성비 쇼핑",
        "best_category": "생활용품 / 의류 / 가전",
        "content_type": "상세페이지 이미지 + 가격 강조",
        "beginner_tip": "🛒 중국 최대 오픈마켓. 판매 시작에 가장 진입장벽 낮음",
        "desc": "중국 최대 오픈마켓. 상세페이지 설득력과 가성비 중심 판매.",
        "color": "#ff6600",
        "style": "促销导向 / 卖点突出 / 价格优势 / 直接转化",
        "hashtag_count": 5,
        "scene_count": "3~5",
        "video_duration": "30~60초",
        "caption_max": 500,
    },
    "douyin": {
        "name_kr": "도우인", "name_cn": "抖音", "icon": "🎵",
        "monthly_users": "월 7억+ 명",
        "main_users": "15~35세 전체",
        "best_category": "패션 / 뷰티 / 식품 / 가전",
        "content_type": "15~60초 숏폼 영상 + 라이브커머스",
        "beginner_tip": "📱 중국판 틱톡. 바이럴 폭발력 최강, 영상 콘텐츠 필수",
        "desc": "중국판 틱톡. 트렌디한 숏폼 영상으로 즉각 구매 유도.",
        "color": "#010101",
        "style": "神曲BGM / 视觉冲击 / 剧情反转 / 快速带货",
        "hashtag_count": 6,
        "scene_count": "4~6",
        "video_duration": "15~60초",
        "caption_max": 300,
    },
    "jingdong": {
        "name_kr": "징동", "name_cn": "京东", "icon": "📦",
        "monthly_users": "월 5억+ 명",
        "main_users": "30~50대, 프리미엄 선호",
        "best_category": "가전 / 디지털 / 브랜드 의류",
        "content_type": "공식 인증 + 스펙 상세 + 빠른 배송",
        "beginner_tip": "🏷️ 프리미엄·공식 브랜드 이미지가 있다면 여기서 신뢰도 확보",
        "desc": "중국 2위 공식 브랜드 쇼핑몰. 정품 신뢰도가 핵심.",
        "color": "#e31837",
        "style": "正品保证 / 参数详细 / 极速物流 / 商务品质",
        "hashtag_count": 5,
        "scene_count": "3~5",
        "video_duration": "30~60초",
        "caption_max": 600,
    },
    "pinduoduo": {
        "name_kr": "핀둬둬", "name_cn": "拼多多", "icon": "🛒",
        "monthly_users": "월 7억+ 명",
        "main_users": "3~4선 도시, 가성비 극단 추구",
        "best_category": "생활용품 / 식품 / 저가 의류",
        "content_type": "공동구매 최저가 + 단순 직관 메시지",
        "beginner_tip": "💰 가성비 제품이라면 폭발적인 물량 판매 가능",
        "desc": "공동구매 플랫폼. '최저가'와 '지인 추천'이 핵심 구매 동기.",
        "color": "#e53935",
        "style": "全网最低 / 拼单裂变 / 简单粗暴 / 厂货直供",
        "hashtag_count": 4,
        "scene_count": "3~4",
        "video_duration": "15~30초",
        "caption_max": 200,
    },
    "tmall": {
        "name_kr": "티몰", "name_cn": "天猫", "icon": "🐱",
        "monthly_users": "월 알리바바 생태계 10억+",
        "main_users": "25~45세 브랜드 가치 중시",
        "best_category": "프리미엄 뷰티 / 패션 / 가전",
        "content_type": "공식 브랜드관 + 프리미엄 이미지",
        "beginner_tip": "⭐ 브랜드 공식 런칭 or 고급화 전략에 최적",
        "desc": "알리바바 프리미엄 브랜드관. 입점 기준 높지만 고급 이미지.",
        "color": "#ff4400",
        "style": "品牌调性 / 高级视觉 / 官方正品 / 尊享服务",
        "hashtag_count": 8,
        "scene_count": "4~6",
        "video_duration": "60~90초",
        "caption_max": 800,
    },
}

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 2. 이미지 테마 (app12.py IMAGE_THEMES 완전 동일)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
NO_TEXT = (
    "ABSOLUTELY NO TEXT, NO LETTERS, NO WORDS, NO FONTS, NO TYPOGRAPHY, "
    "NO LOGOS, NO WATERMARKS, NO LABELS anywhere in the image. "
    "Pure product photography only."
)
PRESERVE_PROMPT_BOOST = (
    "IMPORTANT: Keep the exact product shape, color, texture, and silhouette unchanged. "
    "Only replace/enhance the background environment. "
    "Product must be clearly recognizable and identical to the input image. "
    "Do not alter product form, proportions, or surface details. "
)

IMAGE_THEMES: List[Dict[str, str]] = [
    {
        "name_kr": "☀️ 자연채광 웜톤 거실 컷",
        "desc":    "따뜻한 오후 햇살과 우드 톤의 코지한 거실 무드",
        "style": (
            "Interior lifestyle photography, warm afternoon sunlight streaming through sheer curtains, "
            "natural wood tones and warm beige palette, cozy and inviting living room setting, "
            "soft linen textures, potted plant in the background, "
            "Xiaohongshu interior influencer aesthetic, golden hour warmth. "
        ),
    },
    {
        "name_kr": "🛋️ 모던 미니멀리즘 컷",
        "desc":    "군더더기 없는 화이트/그레이 톤의 세련된 현대 공간",
        "style": (
            "Modern minimalist interior photography, clean white and light grey tones, "
            "sleek contemporary furniture, uncluttered composition with intentional negative space, "
            "soft diffused studio lighting, architectural lines, "
            "Scandinavian-Korean fusion design aesthetic, premium lifestyle feel. "
        ),
    },
    {
        "name_kr": "🌃 무드 시네마틱 라운지 컷",
        "desc":    "은은한 간접 조명과 럭셔리 호텔 라운지 감성",
        "style": (
            "Cinematic luxury lounge interior photography, warm indirect ambient lighting, "
            "deep moody tones with rich shadows, high-end hotel or boutique lounge atmosphere, "
            "plush textures, velvet and marble accents, "
            "sophisticated and aspirational mood, editorial interior style. "
        ),
    },
    {
        "name_kr": "🌿 식물테리어(플랜테리어) 컷",
        "desc":    "반려식물과 자연스럽게 어우러진 프레시한 그린 인테리어",
        "style": (
            "Planteria interior photography, lush green houseplants of various sizes surrounding the furniture, "
            "fresh and natural biophilic design, soft daylight with green leafy shadows, "
            "terracotta pots and natural woven baskets as props, "
            "earthy and refreshing mood, urban jungle Xiaohongshu lifestyle aesthetic. "
        ),
    },
]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 3. 규정 준수
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CATEGORY_BANNED: Dict[str, List[str]] = {
    "건강식품": ["치료", "완치", "의학적 효능", "임상 입증", "FDA 승인"],
    "화장품":   ["주름 제거", "미백 보장", "성형 효과", "피부과 인증"],
    "식품":     ["다이어트 보장", "칼로리 제로 보장", "의사 추천"],
}
BASE_BANNED = ["100%", "最", "立即见效", "治愈", "无副作用", "最好", "第一"]
BANNED_EXPLANATION = {
    "最":       "最 (최고·가장 등 최상급 표현 — 중국 광고법 금지)",
    "100%":     "100% (절대적 수치 보장 — 허위 광고 위험)",
    "立即见效": "立即见效 (즉각 효과 보장 — 과장 광고)",
    "治愈":     "治愈 (치료·완치 의료 주장 — 식품·가구 사용 불가)",
    "无副作用": "无副作用 (부작용 없음 보장 — 의료 효능 주장)",
    "最好":     "最好 (가장 좋음 — 근거 없는 최상급 금지)",
    "第一":     "第一 (1위 표현 — 증빙 없는 순위 금지)",
}

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 4. 데이터 스키마 (app12.py 기반)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
class Brief(BaseModel):
    brand_name:     str
    product_name:   str
    category:       str           = ""
    target:         str           = ""
    price:          Optional[str] = None
    promo:          Optional[str] = None
    key_benefits:   List[str]     = Field(default_factory=list)
    tone:           str           = "真实测评"
    banned_claims:  List[str]     = Field(default_factory=list)
    landing_action: str           = "点进主页领券"

class CreativePack(BaseModel):
    platform:          str = ""
    title_cn:          str = ""
    hook_cn:           str = ""
    body_cn:           str = ""
    cta_cn:            str = ""
    hashtags:          List[str] = Field(default_factory=list)
    thumbnail_text_cn: str = ""
    title_kr:          str = ""
    hook_kr:           str = ""
    body_kr:           str = ""
    cta_kr:            str = ""
    thumbnail_text_kr: str = ""
    subtitles_cn:      List[str] = Field(default_factory=list)
    subtitles_kr:      List[str] = Field(default_factory=list)
    storyboard:        List[Dict[str, str]] = Field(default_factory=list)

# AI 영상 제작 사이트
VIDEO_TOOLS = [
    {
        "key": "runway",    "name": "Runway Gen-4",   "icon": "🎬",
        "url": "https://runwayml.com",
        "price": "월 $15~",
        "desc": "정교한 카메라 워킹·모션 제어. 가장 안정적인 품질. 4K 지원.",
    },
    {
        "key": "kling",     "name": "Kling AI",       "icon": "🐲",
        "url": "https://klingai.com",
        "price": "부분 무료 / 월 ¥66~",
        "desc": "중국 특화 AI. 동양적 텍스처·인물 묘사 최강. 5초~3분 영상.",
    },
    {
        "key": "pika",      "name": "Pika Labs",      "icon": "⚡",
        "url": "https://pika.art",
        "price": "무료(워터마크) / 월 $8~",
        "desc": "초보 친화적 UI. 이미지→영상 변환 기능 탁월. 빠른 생성.",
    },
    {
        "key": "hailuo",    "name": "海螺AI(Hailuo)", "icon": "🌊",
        "url": "https://hailuoai.com",
        "price": "부분 무료",
        "desc": "중국 MiniMax 제작. 사실적 인물 영상·광고 영상에 강점.",
    },
    {
        "key": "luma",      "name": "Luma Dream Machine", "icon": "✨",
        "url": "https://lumalabs.ai/dream-machine",
        "price": "무료 30회/월 / 월 $29.99~",
        "desc": "물리적 사실감 최고 수준. 제품 클로즈업 영상에 특히 우수.",
    },
    {
        "key": "grok",      "name": "Grok (xAI)",         "icon": "🤖",
        "url": "https://grok.com",
        "price": "X Premium 구독 포함 ($8/월~)",
        "desc": "일론 머스크 xAI 모델. 텍스트→이미지·영상 생성. X(구 트위터) 바이럴 연동 가능.",
    },
]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 5. 유틸리티
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def pil_to_b64(img: Image.Image, fmt="PNG") -> str:
    buf = BytesIO(); img.save(buf, format=fmt)
    return base64.b64encode(buf.getvalue()).decode()

def b64_to_pil(b64: str) -> Image.Image:
    return Image.open(BytesIO(base64.b64decode(b64)))

def ensure_list(v) -> List[str]:
    if v is None: return []
    if isinstance(v, list): return [str(x).strip() for x in v if str(x).strip()]
    if isinstance(v, str):
        return [x.strip() for x in v.replace(";", ",").replace("\n", ",").split(",") if x.strip()]
    return [str(v).strip()]

def ensure_str(v) -> str:
    if v is None: return ""
    if isinstance(v, str): return v
    if isinstance(v, list): return "\n".join(str(x) for x in v)
    return str(v)

def ensure_storyboard(v) -> List[Dict[str, str]]:
    if not v or not isinstance(v, list): return []
    return [
        {k: str(val) for k, val in item.items()} if isinstance(item, dict)
        else {"scene": str(item), "duration": "", "visual": "", "caption_cn": "", "caption_kr": ""}
        for item in v
    ]

# CJK 폰트
def _find_cjk_font_path() -> Optional[str]:
    candidates = [
        "C:/Windows/Fonts/malgun.ttf",
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simsun.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/noto-cjk/NotoSansCJKsc-Regular.otf",
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/System/Library/Fonts/PingFang.ttc",
    ]
    for p in candidates:
        if os.path.exists(p): return p
    return None

CJK_FONT_PATH = _find_cjk_font_path()

def _get_font(size: int):
    if CJK_FONT_PATH:
        try: return ImageFont.truetype(CJK_FONT_PATH, size)
        except: pass
    return ImageFont.load_default()

def _draw_wrapped_cjk(draw, text: str, font, max_w: int, start_y: int, fill_color=(255,255,255,255)) -> int:
    """문자 단위 자동 줄바꿈 (CJK 잘림 방지)"""
    lines, cur = [], ""
    for ch in text:
        test = cur + ch
        w = draw.textbbox((0,0), test, font=font)[2]
        if w <= max_w:
            cur = test
        else:
            if cur: lines.append(cur)
            cur = ch
    if cur: lines.append(cur)

    y = start_y
    for line in lines:
        bbox = draw.textbbox((0,0), line, font=font)
        lw = bbox[2]; lh = bbox[3] - bbox[1]
        x = (max_w - lw) // 2
        draw.text((x, y), line, font=font, fill=fill_color, stroke_width=2, stroke_fill=(0,0,0,210))
        y += lh + 10
    return y

# 복사 버튼 (base64 인코딩으로 특수문자 안전)
def copy_btn(text: str, key: str, label="📋 복사", block=False):
    b64 = base64.b64encode(text.encode("utf-8")).decode("ascii")
    w = "width:100%;" if block else ""
    components.html(f"""
    <button id="cb_{key}" data-b64="{b64}"
      onclick="var b=this;var raw=b.getAttribute('data-b64');
               var bytes=Uint8Array.from(atob(raw),c=>c.charCodeAt(0));
               var txt=new TextDecoder('utf-8').decode(bytes);
               navigator.clipboard.writeText(txt).then(()=>{{
                 b.textContent='✅ 복사됨!';
                 setTimeout(()=>b.textContent='{label}',1800);
               }}).catch(()=>{{b.textContent='❌실패';setTimeout(()=>b.textContent='{label}',1800);}});"
      style="font-size:13px;font-weight:600;padding:7px 16px;border:1.5px solid #d1d5db;
             border-radius:8px;background:#fff;cursor:pointer;color:#374151;{w}">{label}</button>
    """, height=42)

# 단계 진행바
def render_step_bar(current: int):
    steps = ["📱 플랫폼", "📝 제품 정보", "🚀 생성 결과", "📊 다운로드"]
    cols = st.columns(len(steps))
    for i, (col, label) in enumerate(zip(cols, steps), 1):
        with col:
            if i < current:
                st.markdown(f"<div style='text-align:center;padding:8px;background:#d1fae5;border-radius:10px;font-size:13px;font-weight:700;color:#065f46;'>✓ {label}</div>", unsafe_allow_html=True)
            elif i == current:
                st.markdown(f"<div style='text-align:center;padding:8px;background:#e53935;border-radius:10px;font-size:13px;font-weight:700;color:#fff;'>▶ {label}</div>", unsafe_allow_html=True)
            else:
                st.markdown(f"<div style='text-align:center;padding:8px;background:#f3f4f6;border-radius:10px;font-size:13px;color:#9ca3af;'>{label}</div>", unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 6. CSS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def apply_css():
    st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: -apple-system, 'Malgun Gothic', 'Microsoft YaHei', sans-serif;
    background: #f5f7fa; color: #1a202c;
}
h1 { font-size:26px!important; font-weight:900!important; color:#1a202c!important; }
h2 { font-size:20px!important; font-weight:800!important; color:#1a202c!important; }
h3 { font-size:17px!important; font-weight:700!important; color:#1a202c!important; }
p, span, li, label, div.stMarkdown p {
    font-size:14px!important; color:#2d3748!important; line-height:1.75!important;
}
/* 카드 */
.card {
    background:#fff; border:1px solid #e2e8f0; border-radius:14px;
    padding:22px 26px; margin-bottom:16px;
    box-shadow:0 2px 8px rgba(0,0,0,0.04);
}
/* 플랫폼 카드 */
.pcard {
    background:#fff; border:2.5px solid #e2e8f0; border-radius:14px;
    padding:18px 14px 14px; text-align:center; cursor:pointer;
    transition:all 0.2s; min-height:200px; margin-bottom:4px;
}
.pcard:hover { border-color:#e53935; transform:translateY(-3px); box-shadow:0 8px 24px rgba(229,57,53,0.14); }
.pcard.sel { border-color:#e53935!important; background:#fff5f5!important; box-shadow:0 0 0 4px rgba(229,57,53,0.16)!important; }
.picon { font-size:38px; display:block; margin-bottom:6px; }
.pcn { font-size:24px; font-weight:900; color:#1a202c; }
.pkr { font-size:13px; color:#6b7280; margin-top:2px; }
.pstat { font-size:11px; font-weight:700; color:#ef4444; margin:5px 0; }
.pwho { font-size:11.5px; color:#374151; background:#f9fafb; border-radius:8px; padding:7px 8px; margin-top:6px; text-align:left; line-height:1.6; }
.ptip { font-size:11px; color:#1d4ed8; font-weight:600; margin-top:6px; }
/* 배지 */
.badge { display:inline-block; padding:3px 10px; border-radius:20px; font-size:12px; font-weight:700; }
.b-red { background:#fef2f2; color:#e53935; border:1px solid #fecaca; }
.b-green { background:#f0fdf4; color:#16a34a; border:1px solid #bbf7d0; }
.b-blue { background:#eff6ff; color:#1d4ed8; border:1px solid #bfdbfe; }
.b-gray { background:#f3f4f6; color:#374151; border:1px solid #d1d5db; }
/* 버튼 */
.stButton>button { border-radius:10px!important; font-weight:600!important; font-size:14px!important; }
.stButton>button[kind="primary"] { background:#e53935!important; border:none!important; color:#fff!important; }
/* 입력 */
.stTextInput input, .stTextArea textarea { font-size:14px!important; color:#1a202c!important; border-radius:8px!important; }
/* 탭 */
.stTabs [data-baseweb="tab"] { font-size:14px!important; font-weight:600!important; }
.stTabs [aria-selected="true"] { color:#e53935!important; border-bottom-color:#e53935!important; }
/* 구분선 */
hr { border-color:#e5e7eb!important; }
/* 알림 */
.stAlert p { font-size:14px!important; }
/* 소형 이미지 컨테이너 */
.img-thumb { border:1px solid #e2e8f0; border-radius:10px; padding:10px; background:#fff; }
/* 텍스트 */
.section-title { font-size:16px; font-weight:700; color:#1a202c; margin:12px 0 8px; }
</style>
""", unsafe_allow_html=True)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 7. 모듈 1 연동 로직
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
MODULE1_JSON_PATH = os.getenv("MODULE1_OUTPUT_PATH", "module1_output.json")

def get_module1_data() -> Dict[str, Any]:
    """3가지 방법으로 모듈1 데이터 수신 (우선순위 순)"""
    # 방법 1: session_state (같은 앱 내 다른 모듈에서 직접 주입)
    if "module1_data" in st.session_state and st.session_state.module1_data:
        return st.session_state.module1_data

    # 방법 2: 공유 JSON 파일
    try:
        if os.path.exists(MODULE1_JSON_PATH):
            with open(MODULE1_JSON_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass

    # 방법 3: URL 파라미터
    try:
        params = st.query_params
        if "brand" in params or "product" in params:
            return {
                "brand_name":   params.get("brand", ""),
                "product_name": params.get("product", ""),
                "category":     params.get("category", ""),
                "price":        params.get("price", ""),
                "promo":        params.get("promo", ""),
                "target":       params.get("target", ""),
                "tone":         params.get("tone", ""),
            }
    except Exception:
        pass
    return {}

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 8. AI 비즈니스 로직
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _call_structured(system: str, user: str, schema: dict, max_retries: int = 2) -> Dict[str, Any]:
    """Structured Outputs with retry + fallback (app12.py 방식)"""
    for attempt in range(max_retries + 1):
        try:
            resp = openai_client.chat.completions.create(
                model=TEXT_MODEL,
                response_format={"type":"json_schema","json_schema":{"name":schema.get("title","output"),"strict":True,"schema":schema}},
                messages=[{"role":"system","content":system},{"role":"user","content":user}],
                temperature=0.85,
            )
            return json.loads(resp.choices[0].message.content)
        except Exception as e:
            if attempt == max_retries:
                try:
                    resp2 = openai_client.chat.completions.create(
                        model=TEXT_MODEL,
                        response_format={"type":"json_object"},
                        messages=[{"role":"system","content":system+"\n\nReturn ONLY valid JSON."},{"role":"user","content":"json\n"+user}],
                        temperature=0.85,
                    )
                    st.warning("⚠️ 구조화 출력 실패 → fallback json_object 사용")
                    return json.loads(resp2.choices[0].message.content)
                except Exception as e2:
                    raise RuntimeError(f"API 호출 실패: {e2}") from e2

# 이미지 분석 스키마
VISION_SCHEMA = {
    "title": "product_analysis",
    "type": "object",
    "properties": {
        "brand_name":    {"type": "string"},
        "product_name":  {"type": "string"},
        "category":      {"type": "string"},
        "target":        {"type": "string"},
        "price_hint":    {"type": "string"},
        "key_benefits":  {"type": "array", "items": {"type": "string"}},
        "tone":          {"type": "string"},
        "color_palette": {"type": "string"},
        "material":      {"type": "string"},
    },
    "required": ["brand_name","product_name","category","target","price_hint","key_benefits","tone","color_palette","material"],
    "additionalProperties": False,
}

def vision_analyze_image(img_bytes: bytes) -> Dict[str, Any]:
    """GPT-4o Vision 이미지 분석 → 한국어 JSON 반환"""
    b64 = base64.b64encode(img_bytes).decode()
    system = (
        "너는 제품 이미지 마케팅 분석 전문가다. "
        "반드시 '한국어'로 모든 필드를 작성한다. "
        "key_benefits는 3~5개 배열. "
        "분석 불가한 필드는 '분석 불가'로 채운다."
    )
    user_content = [
        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}},
        {"type": "text", "text": "위 이미지의 제품을 한국어로 분석해줘. JSON 스키마에 맞게 반환."},
    ]
    try:
        resp = openai_client.chat.completions.create(
            model="gpt-4o",
            response_format={"type":"json_schema","json_schema":{"name":"product_analysis","strict":True,"schema":VISION_SCHEMA}},
            messages=[{"role":"system","content":system},{"role":"user","content":user_content}],
        )
        return json.loads(resp.choices[0].message.content)
    except Exception as e:
        st.warning(f"이미지 분석 오류: {e}")
        return {}

def remove_background(img_bytes: bytes) -> bytes:
    """remove.bg API → rembg 라이브러리 fallback"""
    # 방법 1: remove.bg API
    if REMOVE_BG_KEY:
        try:
            resp = requests.post(
                "https://api.remove.bg/v1.0/removebg",
                files={"image_file": img_bytes},
                data={"size": "auto"},
                headers={"X-Api-Key": REMOVE_BG_KEY},
                timeout=30,
            )
            if resp.status_code == 200:
                return resp.content
        except Exception:
            pass

    # 방법 2: rembg 라이브러리
    if REMBG_AVAILABLE:
        try:
            return rembg_remove(img_bytes)
        except Exception:
            pass

    return img_bytes  # fallback: 원본 반환

BRIEF_SCHEMA = {
    "title": "brief_partial",
    "type": "object",
    "properties": {
        "target":         {"type": "string"},
        "key_benefits":   {"type": "array", "items": {"type": "string"}},
        "tone":           {"type": "string"},
        "banned_claims":  {"type": "array", "items": {"type": "string"}},
        "landing_action": {"type": "string"},
    },
    "required": ["target","key_benefits","tone","banned_claims","landing_action"],
    "additionalProperties": False,
}

@st.cache_data(show_spinner=False)
def build_brief_cached(brand_name, product_name, category, price, promo,
                       extra_notes, tone_style, persona) -> Brief:
    cat_banned = CATEGORY_BANNED.get(category.split("/")[0] if category else "", [])
    all_banned = list(set(BASE_BANNED + cat_banned))
    sys = (
        f"너는 중국 B2C 왕홍 마케팅 전문 기획자다. "
        f"오직 target, key_benefits, tone, banned_claims, landing_action 5개 키만 생성. "
        f"key_benefits: 3~5개 배열. banned_claims에 반드시 포함: {all_banned}. "
        f"tone 스타일: {tone_style or '真实测评'}. "
        f"타겟 페르소나: {persona or '자동 추론'}. "
        f"반드시 한국어로 작성."
    )
    user = f"브랜드: {brand_name}\n제품: {product_name}\n카테고리: {category}\n가격: {price}\n프로모션: {promo}\n추가메모: {extra_notes}"
    data = _call_structured(sys, user, BRIEF_SCHEMA)
    data["key_benefits"]  = ensure_list(data.get("key_benefits"))
    data["banned_claims"] = list(set(ensure_list(data.get("banned_claims")) + all_banned))
    return Brief(
        brand_name=brand_name or "미입력", product_name=product_name,
        category=category or "미입력", price=price or None, promo=promo or None,
        target=data.get("target",""), key_benefits=data["key_benefits"],
        tone=data.get("tone", tone_style or "真实测评"),
        banned_claims=data["banned_claims"],
        landing_action=data.get("landing_action","点进主页领券"),
    )

CREATIVE_PACK_SCHEMA = {
    "title": "creative_pack",
    "type": "object",
    "properties": {
        "platform":          {"type":"string"},
        "title_cn":          {"type":"string"}, "hook_cn":  {"type":"string"},
        "body_cn":           {"type":"string"}, "cta_cn":   {"type":"string"},
        "hashtags":          {"type":"array","items":{"type":"string"}},
        "thumbnail_text_cn": {"type":"string"}, "title_kr": {"type":"string"},
        "hook_kr":           {"type":"string"}, "body_kr":  {"type":"string"},
        "cta_kr":            {"type":"string"}, "thumbnail_text_kr":{"type":"string"},
        "subtitles_cn":      {"type":"array","items":{"type":"string"}},
        "subtitles_kr":      {"type":"array","items":{"type":"string"}},
        "storyboard": {
            "type":"array","items":{
                "type":"object",
                "properties":{
                    "scene":{"type":"string"},"duration":{"type":"string"},
                    "visual":{"type":"string"},"caption_cn":{"type":"string"},
                    "caption_kr":{"type":"string"},
                },
                "required":["scene","duration","visual","caption_cn","caption_kr"],
                "additionalProperties":False,
            },
        },
    },
    "required":["platform","title_cn","hook_cn","body_cn","cta_cn","hashtags",
                "thumbnail_text_cn","title_kr","hook_kr","body_kr","cta_kr",
                "thumbnail_text_kr","subtitles_cn","subtitles_kr","storyboard"],
    "additionalProperties":False,
}

@st.cache_data(show_spinner=False)
def build_creative_pack_cached(brief_json: str, platform_key: str, tone_style: str) -> CreativePack:
    brief = Brief.model_validate_json(brief_json)
    p = PLATFORMS[platform_key]
    sys = (
        f"너는 중국 {p['name_kr']} ({p['name_cn']}) 전문 마케터다.\n"
        f"플랫폼 스타일: {p['style']}\n"
        f"권장 영상: {p['video_duration']}\n"
        f"해시태그: 정확히 {p['hashtag_count']}개 (#포함)\n"
        f"본문: {p['caption_max']}자 이내\n"
        f"subtitles_cn: 숏폼 자막 8줄 이내, 한 줄 최대 18자\n"
        f"storyboard: {p['scene_count']}개 씬\n"
        f"콘텐츠 톤: {tone_style or brief.tone}\n"
        f"절대 금지 표현: {brief.banned_claims}\n"
        f"_cn → 자연스러운 중국어 / _kr → 정확한 한국어 번역"
    )
    user = json.dumps({"brief": brief.model_dump(), "platform": platform_key}, ensure_ascii=False)
    data = _call_structured(sys, user, CREATIVE_PACK_SCHEMA)

    str_fields = ["platform","title_cn","hook_cn","body_cn","cta_cn","thumbnail_text_cn",
                  "title_kr","hook_kr","body_kr","cta_kr","thumbnail_text_kr"]
    for f in str_fields:
        data[f] = ensure_str(data.get(f))
    data["hashtags"]     = ensure_list(data.get("hashtags"))[:p["hashtag_count"]]
    data["subtitles_cn"] = ensure_list(data.get("subtitles_cn"))
    data["subtitles_kr"] = ensure_list(data.get("subtitles_kr"))
    data["storyboard"]   = ensure_storyboard(data.get("storyboard"))
    data.setdefault("platform", platform_key)
    pack = CreativePack(**data)

    # 컴플라이언스 검수 → 위반 시 1회 재생성
    violations = _check_compliance(pack, brief.banned_claims)
    if violations:
        sys2 = sys + f"\n\n⚠️ 금지 표현 제거 후 재생성: {violations}"
        try:
            data2 = _call_structured(sys2, user, CREATIVE_PACK_SCHEMA)
            for f in str_fields: data2[f] = ensure_str(data2.get(f))
            data2["hashtags"]     = ensure_list(data2.get("hashtags"))[:p["hashtag_count"]]
            data2["subtitles_cn"] = ensure_list(data2.get("subtitles_cn"))
            data2["subtitles_kr"] = ensure_list(data2.get("subtitles_kr"))
            data2["storyboard"]   = ensure_storyboard(data2.get("storyboard"))
            data2.setdefault("platform", platform_key)
            pack2 = CreativePack(**data2)
            violations = _check_compliance(pack2, brief.banned_claims)
            pack = pack2
        except Exception:
            pass

    pack._compliance_violations = violations
    return pack

def _check_compliance(pack: CreativePack, banned: List[str]) -> List[str]:
    texts = [pack.title_cn, pack.hook_cn, pack.body_cn, pack.cta_cn,
             pack.title_kr, pack.hook_kr, pack.body_kr] + pack.subtitles_cn
    found = []
    for b in banned:
        for t in texts:
            if b.lower() in t.lower() and b not in found:
                found.append(b)
    return found

def retranslate_body_cn(body_kr: str, brief: Brief, platform_key: str) -> str:
    p = PLATFORMS[platform_key]
    resp = openai_client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role":"system","content":(
                f"너는 한→중 마케팅 번역 전문가다. {p['name_cn']} ({p['name_kr']}) 스타일에 맞는 자연스러운 중국어로 번역. "
                f"직역 금지. 중국 MZ세대 공감 표현 사용. 금지 표현: {', '.join(brief.banned_claims)}"
            )},
            {"role":"user","content":f"아래 한국어 마케팅 본문을 {p['name_cn']} 스타일 중국어로 번역:\n\n{body_kr}"},
        ],
        temperature=0.7, max_tokens=600,
    )
    return resp.choices[0].message.content.strip()

# DeepSeek 문구/해시태그 추천
def generate_deepseek_copies(brief: Brief, platform_key: str) -> Dict[str, Any]:
    """DeepSeek로 이미지별 문구, 해시태그 추천 (한국어 번역 포함)"""
    p = PLATFORMS[platform_key]
    sys_p = (
        "너는 중국 왕홍 마케팅 카피라이터다. 이미지에 합성할 짧고 강렬한 문구와 해시태그를 생성한다. "
        "반드시 JSON으로 반환한다."
    )
    user_p = (
        f"플랫폼: {p['name_cn']} ({p['name_kr']})\n"
        f"제품: {brief.product_name} / 브랜드: {brief.brand_name}\n"
        f"타겟: {brief.target}\n핵심 강점: {', '.join(brief.key_benefits[:3])}\n"
        f"금지어: {brief.banned_claims[:5]}\n\n"
        "아래 JSON 형식으로 반환:\n"
        '{"overlays": [{"zh":"문구(10자이내)","kr":"한국어번역"},{"zh":"","kr":""},{"zh":"","kr":""},{"zh":"","kr":""}], '
        '"hashtags_cn":["#해시태그1","#해시태그2",...10개], '
        '"hashtags_kr":["#번역1","#번역2",...10개]}'
    )
    try:
        resp = deepseek_client.chat.completions.create(
            model="deepseek-chat",
            messages=[{"role":"system","content":sys_p},{"role":"user","content":user_p}],
            response_format={"type":"json_object"},
        )
        return json.loads(resp.choices[0].message.content)
    except Exception as e:
        st.warning(f"DeepSeek 추천 오류 (기본값 사용): {e}")
        return {
            "overlays": [
                {"zh":"韩国爆款来了","kr":"한국 히트 상품 등장"},
                {"zh":"精选品质之选","kr":"엄선된 프리미엄 품질"},
                {"zh":"生活高级感","kr":"일상에 고급스러움을"},
                {"zh":"现在购买更优惠","kr":"지금 구매 시 특가"},
            ],
            "hashtags_cn": [f"#{brief.product_name}","#韩国好物","#种草","#生活方式","#家居","#精选","#好物推荐","#品质生活","#韩国品牌","#必买"],
            "hashtags_kr": [f"#{brief.product_name}","#한국직구","#추천","#라이프스타일","#홈데코","#셀렉션","#강추","#품질","#한국브랜드","#필수템"],
        }

# 이미지 생성 (app12.py _call_edit_api 완전 동일)
def _call_edit_api(img_rgba: Image.Image, prompt: str, size: str) -> Optional[Image.Image]:
    buf = BytesIO()
    img_rgba.save(buf, format="PNG")
    buf.seek(0)
    try:
        result = openai_client.images.edit(
            model=IMAGE_MODEL,
            image=("product.png", buf, "image/png"),
            prompt=prompt, n=1, size=size,
        )
        b64 = result.data[0].b64_json
        if b64: return b64_to_pil(b64)
        url = result.data[0].url
        if url:
            import urllib.request
            with urllib.request.urlopen(url) as r:
                return Image.open(BytesIO(r.read()))
    except Exception as e:
        raise e
    return None

@st.cache_data(show_spinner=False)
def generate_images_cached(
    img_b64: Optional[str],
    brief_json: str,
    image_size: str,
    synth_mode: str,
    fast_mode: bool,
) -> List[Tuple[str, str, Optional[str]]]:
    """app12.py _generate_images_cached 완전 동일 로직"""
    brief = Brief.model_validate_json(brief_json)
    use_rembg = (synth_mode == "preserve") and REMBG_AVAILABLE
    themes = IMAGE_THEMES[:2] if fast_mode else IMAGE_THEMES

    results = []
    if img_b64:
        product_img = b64_to_pil(img_b64)
        if use_rembg:
            try:
                buf_in = BytesIO(); product_img.save(buf_in, format="PNG")
                removed = rembg_remove(buf_in.getvalue())
                base_img = Image.open(BytesIO(removed)).convert("RGBA")
            except Exception:
                base_img = product_img.convert("RGBA")
        else:
            base_img = product_img.convert("RGBA")

        for theme in themes:
            preserve_boost = PRESERVE_PROMPT_BOOST if synth_mode == "preserve" else ""
            prompt = (
                f"{preserve_boost}{theme['style']}"
                f"Hero product: '{brief.product_name}' by brand '{brief.brand_name}'. "
                f"Target customer: {brief.target}. "
                f"Key benefits (show visually, not as text): {', '.join(brief.key_benefits[:2])}. "
                f"No medical claims. No before/after comparison. {NO_TEXT}"
            )
            try:
                img = _call_edit_api(base_img, prompt, image_size)
                results.append((theme["name_kr"], theme["desc"], pil_to_b64(img) if img else None))
            except Exception as e:
                st.warning(f"⚠️ [{theme['name_kr']}] 생성 실패: {e}")
                results.append((theme["name_kr"], theme["desc"], None))
    else:
        # 텍스트 프롬프트 기반 생성
        for theme in themes:
            prompt = (
                f"{theme['style']}"
                f"Product: '{brief.product_name}' by brand '{brief.brand_name}'. "
                f"Target customer: {brief.target}. Xiaohongshu Chinese e-commerce style. {NO_TEXT}"
            )
            kwargs: Dict[str, Any] = {"model": IMAGE_MODEL, "prompt": prompt, "size": image_size, "n": 1}
            if "dall-e" in IMAGE_MODEL: kwargs["response_format"] = "b64_json"
            try:
                result = openai_client.images.generate(**kwargs)
                b64 = result.data[0].b64_json
                results.append((theme["name_kr"], theme["desc"], b64 if b64 else None))
            except Exception as e:
                st.warning(f"⚠️ [{theme['name_kr']}] 생성 실패: {e}")
                results.append((theme["name_kr"], theme["desc"], None))

    return results

# 문구 PIL 합성
def overlay_text_on_image(base_b64: str, text_zh: str) -> str:
    """반투명 하단 배너에 중국어 문구 자동 줄바꿈 합성"""
    img = b64_to_pil(base_b64).convert("RGBA")
    W, H = img.size
    font_size = max(24, int(W * 0.048))
    font = _get_font(font_size)
    pad = int(W * 0.08); max_w = W - pad * 2

    # 줄바꿈 예상 높이 계산
    tmp_draw = ImageDraw.Draw(Image.new("RGBA", (W, H)))
    end_y = _draw_wrapped_cjk(tmp_draw, text_zh, font, max_w, 0)
    banner_h = end_y + 60

    overlay = Image.new("RGBA", (W, banner_h), (0, 0, 0, 165))
    img.paste(overlay, (0, H - banner_h), overlay)

    draw = ImageDraw.Draw(img)
    _draw_wrapped_cjk(draw, text_zh, font, max_w, H - banner_h + 30)

    return pil_to_b64(img.convert("RGB"))

# app12.py _overlay_text_on_image (브리프 정보 오버레이)
def overlay_brief_on_image(img: Image.Image, brief: Brief, pack: Optional[CreativePack] = None) -> Image.Image:
    img = img.copy().convert("RGBA")
    W, H = img.size
    banner_h = int(H * 0.28)
    overlay = Image.new("RGBA", (W, banner_h), (0, 0, 0, 155))
    img.paste(overlay, (0, H - banner_h), overlay)

    title_f = _get_font(int(W * 0.055))
    body_f  = _get_font(int(W * 0.038))
    small_f = _get_font(int(W * 0.030))
    draw = ImageDraw.Draw(img)
    y = H - banner_h + int(H * 0.015)
    pad = int(W * 0.04)

    draw.text((pad, y), brief.product_name, font=title_f, fill=(255,255,255,255))
    y += int(W * 0.065)
    for b in brief.key_benefits[:2]:
        draw.text((pad, y), f"✓ {b[:20]}", font=body_f, fill=(255,220,100,255))
        y += int(W * 0.045)
    if brief.price:
        price_text = f"💰 {brief.price}"
        if brief.promo: price_text += f"  🎁 {brief.promo[:15]}"
        draw.text((pad, y), price_text, font=body_f, fill=(255,180,180,255))
        y += int(W * 0.045)
    cta = pack.cta_cn[:20] if pack and pack.cta_cn else "点进主页领券 →"
    draw.text((pad, y), cta, font=small_f, fill=(180,230,255,255))
    return img.convert("RGB")

# 왕홍 제안서
PROPOSAL_SCHEMA = {
    "title": "proposal",
    "type": "object",
    "properties": {
        "dm_short_kr":{"type":"string"},"dm_short_cn":{"type":"string"},
        "dm_long_kr": {"type":"string"},"dm_long_cn": {"type":"string"},
        "email_subject_kr":{"type":"string"},"email_subject_cn":{"type":"string"},
        "email_body_kr":   {"type":"string"},"email_body_cn":   {"type":"string"},
    },
    "required":["dm_short_kr","dm_short_cn","dm_long_kr","dm_long_cn",
                "email_subject_kr","email_subject_cn","email_body_kr","email_body_cn"],
    "additionalProperties":False,
}

def generate_wanghong_proposal(brief: Brief, platform_key: str,
                                collab_count="1~2개", collab_format="图文+Reel",
                                collab_deadline="협의 후 결정", collab_benefit="제품 무료 제공 + 커미션",
                                wanghong_name: str = "", wanghong_tier: str = "", wanghong_style: str = "") -> dict:
    p = PLATFORMS[platform_key]
    # 왕홍 이름이 있으면 맞춤 제안서 생성
    name_line = f"수신 왕홍 이름/닉네임: {wanghong_name}" if wanghong_name else "수신 왕홍: (이름 미입력 — 일반 제안서 형식)"
    tier_line = f"왕홍 티어: {wanghong_tier}" if wanghong_tier else ""
    style_line = f"왕홍 콘텐츠 스타일: {wanghong_style}" if wanghong_style else ""
    sys = (
        f"너는 한중 마케팅 협업 전문 에이전트다. {p['name_cn']} 왕홍에게 보내는 3종 협업 제안서를 생성한다.\n"
        "- dm_short: 100~150자 짧은 DM / dm_long: 200~300자 / email: 400~700자\n"
        "- 각각 한국어(_kr)·중국어(_cn) 쌍으로 생성\n"
        "- 정중하고 전문적인 비즈니스 문체. 브랜드명·제품명 변경 금지.\n"
        + (f"- 왕홍 이름({wanghong_name})을 제안서 첫 부분에 반드시 포함: 예) '{wanghong_name}님 안녕하세요' / '亲爱的{wanghong_name}'\n" if wanghong_name else "")
        + (f"- 해당 티어({wanghong_tier})에 맞는 제안 금액 수준과 협업 방식 제안\n" if wanghong_tier else "")
        + (f"- 왕홍의 콘텐츠 스타일({wanghong_style})이 브랜드와 잘 맞는다는 내용 포함\n" if wanghong_style else "")
    )
    user = (
        f"브랜드: {brief.brand_name} / 제품: {brief.product_name} ({brief.category})\n"
        f"타겟: {brief.target} / 강점: {', '.join(brief.key_benefits[:3])}\n"
        f"가격: {brief.price or '미입력'} / 프로모션: {brief.promo or '없음'}\n"
        f"협업 조건: 콘텐츠 {collab_count}개 / 형식: {collab_format} / 납기: {collab_deadline} / 혜택: {collab_benefit}\n"
        f"{name_line}\n{tier_line}\n{style_line}\n"
        "위 정보로 3종 협업 제안서 생성."
    )
    return _call_structured(sys, user, PROPOSAL_SCHEMA)

# 엑셀 리포트
def generate_excel_report(brief: Brief, pack: CreativePack, proposals: dict,
                           themed_meta: List[Tuple[str,str,Any]], platform_key: str) -> bytes:
    wb = Workbook(); ws = wb.active
    ws.title = "왕홍 마케팅 기획안"
    RED = "C0392B"; DKRED = "922B21"; LTRED = "FADBD8"; WHITE = "FFFFFF"

    def hf(bold=True, color=WHITE, size=11): return XLFont(name="Arial", bold=bold, color=color, size=size)
    def bf(color="1A1A2E", size=10): return XLFont(name="Arial", color=color, size=size)
    def fill(h): return PatternFill("solid", start_color=h, fgColor=h)
    def al(wrap=True): return Alignment(horizontal="left", vertical="center", wrap_text=wrap)
    def ac(wrap=True): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    thin = Side(style="thin", color="DDDDDD")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 65
    p_name = PLATFORMS[platform_key]['name_kr']

    ws.merge_cells("A1:C1")
    ws["A1"] = f"🇨🇳 왕홍 마케팅 기획안 — {brief.product_name} ({brief.brand_name}) | {p_name}"
    ws["A1"].font = XLFont(name="Arial", bold=True, color=WHITE, size=14)
    ws["A1"].fill = fill(RED); ws["A1"].alignment = ac(); ws["A1"].border = bd
    ws.row_dimensions[1].height = 32

    def add(cat, sub, val, first=False):
        r = ws.max_row + 1
        ws.row_dimensions[r].height = max(16, min(120, 16 + str(val).count("\n") * 14))
        c1 = ws.cell(r,1,cat); c2 = ws.cell(r,2,sub); c3 = ws.cell(r,3,str(val))
        c1.font = hf(bold=first, color=WHITE if first else DKRED, size=11 if first else 10)
        c1.fill = fill(DKRED if first else LTRED); c1.alignment = ac(); c1.border = bd
        c2.font = hf(bold=True, color=DKRED, size=10)
        c2.fill = fill(LTRED); c2.alignment = al(); c2.border = bd
        c3.font = bf(); c3.fill = fill(WHITE); c3.alignment = al(); c3.border = bd

    add("1. 캠페인 개요","브랜드명",brief.brand_name,True)
    add("1. 캠페인 개요","제품명",brief.product_name)
    add("1. 캠페인 개요","카테고리",brief.category)
    add("1. 캠페인 개요","타겟 고객",brief.target)
    add("1. 캠페인 개요","핵심 소구점","\n".join(f"• {b}" for b in brief.key_benefits))
    add("1. 캠페인 개요","가격",brief.price or "미입력")
    add("1. 캠페인 개요","마케팅 톤",brief.tone)

    add("2. 이미지 전략","생성 테마 수",f"{len([x for _,_,x in themed_meta if x])}개 성공 / {len(themed_meta)}개 시도",True)
    for n, d, _ in themed_meta:
        add("2. 이미지 전략", n, d)

    add(f"3. {p_name} 콘텐츠","제목(CN)",pack.title_cn,True)
    add(f"3. {p_name} 콘텐츠","제목(KR)",pack.title_kr)
    add(f"3. {p_name} 콘텐츠","본문(CN)",pack.body_cn)
    add(f"3. {p_name} 콘텐츠","본문(KR)",pack.body_kr)
    add(f"3. {p_name} 콘텐츠","CTA(CN)",pack.cta_cn)
    add(f"3. {p_name} 콘텐츠","해시태그"," ".join(pack.hashtags))

    add("4. 영상 기획","숏폼 자막(CN)","\n".join(f"{i}. {s}" for i,s in enumerate(pack.subtitles_cn,1)),True)
    add("4. 영상 기획","숏폼 자막(KR)","\n".join(f"{i}. {s}" for i,s in enumerate(pack.subtitles_kr,1)))
    for sc in pack.storyboard:
        add("4. 영상 기획","스토리보드",
            f"[{sc.get('scene','')}] {sc.get('duration','')}\n화면: {sc.get('visual','')}\nCN: {sc.get('caption_cn','')}\nKR: {sc.get('caption_kr','')}")

    add("5. 왕홍 제안서","DM 짧은버전(KR)",proposals.get("dm_short_kr","미생성"),True)
    add("5. 왕홍 제안서","DM 짧은버전(CN)",proposals.get("dm_short_cn","미생성"))
    add("5. 왕홍 제안서","이메일(KR)",proposals.get("email_body_kr","미생성"))
    add("5. 왕홍 제안서","이메일(CN)",proposals.get("email_body_cn","미생성"))

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 9. 메인 UI
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def init_session():
    defaults = {
        "step": 1, "data": {},
        "magic_prompt": "", "extra_notes": "",
        "wanghong_proposals": {},
        "is_generated": False,
        "ss_brief": None, "ss_pack": None,
        "ss_themed_raw": [], "ss_use_rembg": False,
        "ss_uploaded": False, "ss_deepseek": {},
        "retranslated": {},
        # 이미지별 문구 합성 state
        "overlay_state": {},
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

def main():
    st.set_page_config(page_title="🇨🇳 중국 왕홍 마케팅 자동 생성기 v7", layout="wide", initial_sidebar_state="expanded")
    apply_css()
    init_session()

    # ── 사이드바 ──────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("## ⚙️ 설정")

        image_size_label = st.selectbox("📐 이미지 규격", list(IMAGE_SIZE_OPTIONS.keys()), index=0)
        image_size = IMAGE_SIZE_OPTIONS[image_size_label]

        st.divider()
        speed_mode = st.radio("⚡ 생성 모드", ["⚡ 빠른 모드 (이미지 2장)", "💎 고퀄 모드 (이미지 4장)"], index=1)
        fast_mode = "빠른" in speed_mode

        overlay_on = st.toggle("🖊️ 브리프 정보 오버레이 (브랜드/가격/CTA)", value=False)
        if overlay_on and not CJK_FONT_PATH:
            st.warning("⚠️ CJK 폰트 미발견 — 기본 폰트 사용")

        st.divider()
        st.markdown("**🎯 타겟 고객**")
        # _autofill_persona: 이미지 분석 결과를 1회만 채운 뒤 삭제 (사용자 재입력 방해 방지)
        _persona_default = st.session_state.pop("_autofill_persona", st.session_state.get("_saved_persona", ""))
        st.session_state["_saved_persona"] = _persona_default
        persona = st.text_input("타겟 고객 입력", value=_persona_default, placeholder="예: 신혼부부, 30대 인테리어 관심 여성", label_visibility="collapsed")
        if persona != _persona_default:
            st.session_state["_saved_persona"] = persona

        st.markdown("**🎨 원하는 마케팅 느낌**")
        _tone_default = st.session_state.pop("_autofill_tone", st.session_state.get("_saved_tone", ""))
        st.session_state["_saved_tone"] = _tone_default
        tone_style = st.text_input("마케팅 톤 입력", value=_tone_default, placeholder="예: 감성적이고 고급스러운, 친근하고 일상적인", label_visibility="collapsed")
        if tone_style != _tone_default:
            st.session_state["_saved_tone"] = tone_style
        st.caption("💡 '내돈내산 후기처럼', '프리미엄 럭셔리' 등 자유롭게 입력하세요.")

        st.divider()
        st.caption(f"TEXT: `{TEXT_MODEL}` | IMAGE: `{IMAGE_MODEL}`")
        if REMBG_AVAILABLE: st.success("✅ rembg 원본 보존 모드 활성")
        else: st.info("ℹ️ rembg 미설치 — 프롬프트 강화 모드")
        if REMOVE_BG_KEY: st.success("✅ remove.bg API 연결됨")

    # ── 헤더 ──────────────────────────────────────────────────────
    st.markdown("""
    <div style="padding:20px 0 4px;">
      <div style="font-size:28px;font-weight:900;color:#1a202c;">🇨🇳 중국 왕홍 마케팅 자동 생성기</div>
      <div style="font-size:14px;color:#6b7280;margin-top:4px;">
        플랫폼 선택 → 제품 업로드 & 정보 입력 → AI 이미지·카피·영상 프롬프트 자동 생성
      </div>
    </div>
    """, unsafe_allow_html=True)
    st.divider()

    step = st.session_state.step
    render_step_bar(step)

    # ══════════════════════════════════════════════════════════════
    # STEP 1: 플랫폼 선택
    # ══════════════════════════════════════════════════════════════
    if step == 1:
        st.markdown("## 📱 STEP 1 — 진출 플랫폼 선택")
        st.markdown("""
        <div style="background:#eff6ff;border-left:4px solid #3b82f6;padding:14px 18px;border-radius:8px;margin-bottom:20px;">
        💡 <strong>중국 플랫폼이 처음이신가요?</strong> 아래 카드에서 이용자 수·주요 고객·잘 팔리는 방식을 확인하고 선택하세요.<br>
        선택한 플랫폼의 특성에 맞춰 <strong>이미지·카피·해시태그·영상 프롬프트가 자동으로 최적화</strong>됩니다.
        </div>
        """, unsafe_allow_html=True)

        plat_keys = list(PLATFORMS.keys())
        rows = [plat_keys[:3], plat_keys[3:]]
        for row in rows:
            cols = st.columns(3, gap="medium")
            for col, k in zip(cols, row):
                p = PLATFORMS[k]
                is_sel = st.session_state.data.get("platform") == k
                sel_cls = "sel" if is_sel else ""
                chk = "✅ " if is_sel else ""
                with col:
                    st.markdown(f"""
                    <div class="pcard {sel_cls}">
                      <span class="picon">{p['icon']}</span>
                      <div class="pcn">{p['name_cn']}</div>
                      <div class="pkr">{chk}{p['name_kr']}</div>
                      <div class="pstat">👥 {p['monthly_users']}</div>
                      <div class="pwho">
                        🎯 <strong>주 이용자:</strong> {p['main_users']}<br>
                        🏷️ <strong>잘 팔리는 것:</strong> {p['best_category']}<br>
                        📌 <strong>콘텐츠 형식:</strong> {p['content_type']}
                      </div>
                      <div class="ptip">{p['beginner_tip']}</div>
                    </div>
                    """, unsafe_allow_html=True)
                    btn_label = "✅ 선택됨" if is_sel else f"이 플랫폼 선택"
                    if st.button(btn_label, key=f"pb_{k}", use_container_width=True,
                                  type="primary" if is_sel else "secondary"):
                        st.session_state.data["platform"] = k
                        st.rerun()

        sel_k = st.session_state.data.get("platform")
        if sel_k:
            p = PLATFORMS[sel_k]
            st.success(f"✅ **{p['name_cn']} ({p['name_kr']})** 선택됨 — {p['desc']}")
            st.markdown("**📌 이 콘텐츠의 주요 사용처**")
            st.session_state.data["purpose"] = st.radio(
                "용도", ["메인 대표 이미지 (썸네일)", "상세페이지 삽입용", "SNS 피드용 (샤오홍슈/도우인)"],
                horizontal=True, label_visibility="collapsed",
            )
            st.markdown("")
            if st.button("다음 단계 →  제품 정보 입력", type="primary", use_container_width=True):
                st.session_state.step = 2
                st.rerun()
        else:
            st.info("⬆️ 위에서 진출할 플랫폼을 선택해 주세요.")

    # ══════════════════════════════════════════════════════════════
    # STEP 2: 제품 정보 입력 & 이미지 업로드 & 생성
    # ══════════════════════════════════════════════════════════════
    elif step == 2:
        sel_k = st.session_state.data.get("platform", "xiaohongshu")
        p_info = PLATFORMS[sel_k]
        st.markdown(f"## 📝 STEP 2 — 제품 정보 입력")
        st.markdown(f'<span class="badge b-red">📱 {p_info["name_cn"]} ({p_info["name_kr"]})</span> <span class="badge b-gray">📌 {st.session_state.data.get("purpose","")}</span>', unsafe_allow_html=True)

        # 모듈1 데이터 자동 로드
        m1 = get_module1_data()
        if m1:
            st.info(f"✅ 모듈1에서 제품 데이터를 불러왔습니다: **{m1.get('product_name','')}**")

        inp_a, inp_b = st.columns([1, 1], gap="large")

        with inp_a:
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown("#### 📦 제품 기본 정보")

            # ★ 자동 채우기 핵심 로직:
            # Streamlit 제약: 위젯이 렌더링된 후에는 같은 key의 session_state 수정 불가.
            # 해결책: vision_analysis 결과를 "_pending_autofill" 플래그로 저장해두고,
            # 위젯 렌더링 직전(이 위치)에 위젯 key에 반영한 뒤 플래그 제거.
            if st.session_state.get("_pending_autofill"):
                va = st.session_state["_pending_autofill"]
                if va.get("brand_name")   and va["brand_name"]   != "분석 불가": st.session_state["in_brand"]   = va["brand_name"]
                if va.get("product_name") and va["product_name"] != "분석 불가": st.session_state["in_product"] = va["product_name"]
                if va.get("category")     and va["category"]     != "분석 불가": st.session_state["in_cat"]     = va["category"]
                if va.get("price_hint")   and va["price_hint"]   != "분석 불가": st.session_state["in_price"]   = va["price_hint"]
                # persona/tone은 사이드바 위젯 key를 쓰지 않으므로 별도 키에 저장 후 value= 로 반영
                if va.get("target") and va["target"] != "분석 불가":
                    st.session_state["_autofill_persona"] = va["target"]
                if va.get("tone") and va["tone"] != "분석 불가":
                    st.session_state["_autofill_tone"] = va["tone"]
                del st.session_state["_pending_autofill"]

            # 초기값 세팅 (모듈1 데이터 또는 공란)
            if "in_brand"   not in st.session_state: st.session_state["in_brand"]   = m1.get("brand_name", "")
            if "in_product" not in st.session_state: st.session_state["in_product"] = m1.get("product_name", "")
            if "in_cat"     not in st.session_state: st.session_state["in_cat"]     = m1.get("category", "")
            if "in_price"   not in st.session_state: st.session_state["in_price"]   = m1.get("price", "")
            if "in_promo"   not in st.session_state: st.session_state["in_promo"]   = m1.get("promo", "")

            a1, a2 = st.columns(2)
            with a1: input_brand   = st.text_input("브랜드명 *", placeholder="예: 비타제주", key="in_brand")
            with a2: input_product = st.text_input("제품명 *",   placeholder="예: 라탄 소파 2인용", key="in_product")
            input_category = st.text_input("카테고리", placeholder="예: 가구/인테리어, 건강식품, 화장품", key="in_cat")
            a3, a4 = st.columns(2)
            with a3: input_price = st.text_input("가격 (위안/원)", placeholder="예: 69위안 / 35,000원", key="in_price")
            with a4: input_promo = st.text_input("프로모션",       placeholder="예: 출시 기념 20% 할인", key="in_promo")

            # AI 마케팅 포인트 자동 추천
            if st.button("✨ AI 마케팅 포인트 자동 추천", use_container_width=True):
                if not input_product.strip():
                    st.warning("제품명을 먼저 입력해주세요.")
                else:
                    with st.spinner("AI가 마케팅 포인트를 분석 중..."):
                        try:
                            resp = openai_client.chat.completions.create(
                                model="gpt-4o",
                                messages=[
                                    {"role":"system","content":(
                                        f"너는 중국 {p_info['name_kr']} ({p_info['name_cn']}) 마케팅 전략 수석 기획자다. "
                                        f"플랫폼 스타일: {p_info['style']}. "
                                        f"타겟: {persona or '자동추론'}. "
                                        f"마케팅 톤: {tone_style or '진정성 있는 후기 스타일'}. "
                                        "반드시 한국어로 작성. "
                                        "각 항목을 최소 3~5문장 이상 구체적으로 분석한다. "
                                        "실제 왕홍 마케팅 현장에서 사용하는 전략과 트렌드를 반영한다."
                                    )},
                                    {"role":"user","content":(
                                        f"브랜드:{input_brand}\n제품:{input_product}\n카테고리:{input_category}\n"
                                        f"가격:{input_price}\n프로모션:{input_promo}\n"
                                        f"타겟:{persona or '자동추론'}\n마케팅느낌:{tone_style or '자동추론'}\n\n"
                                        "아래 7가지 항목으로 상세한 마케팅 전략 명세서를 작성해줘. "
                                        "각 항목은 **볼드 제목** + 3~5문장 이상의 구체적 내용으로 작성:\n\n"
                                        "**1. 핵심 소구점 (USP)** — 이 타겟 고객에게 가장 강력하게 와닿는 강점 3가지를 구체적으로 설명. "
                                        "경쟁 제품 대비 차별점 포함.\n\n"
                                        "**2. 타겟 고객 페르소나 분석** — 나이, 성별, 소득, 라이프스타일, 구매 습관, 주요 고민사항을 상세히 묘사. "
                                        "이 고객이 이 제품을 구매하게 되는 심리적 동기 포함.\n\n"
                                        "**3. 페인포인트 & 솔루션** — 타겟 고객의 구체적인 3가지 고민과 이 제품이 각각 어떻게 해결해주는지 서술.\n\n"
                                        "**4. 시각적 무드 & 콘텐츠 방향** — 이미지/영상 컨셉, 색감, 배경 스타일, 소품 아이디어. "
                                        f"{p_info['name_cn']} 플랫폼에 최적화된 비주얼 전략.\n\n"
                                        "**5. 추천 콘텐츠 톤앤매너** — 문체 방향 (1인칭/종초/리뷰 등), 감정적 어필 포인트, "
                                        "중국 MZ세대가 공감할 표현 방식 제안.\n\n"
                                        "**6. 가격·프로모션 전략** — 현재 가격/프로모션을 활용한 구매 유도 전략. "
                                        "중국 플랫폼 특성에 맞는 할인·번들·한정 구성 제안.\n\n"
                                        "**7. 예상 KPI & 바이럴 전략** — 조회수, 저장수, 전환율 목표 수치 제안. "
                                        "해시태그 전략, 왕홍 티어 추천 (나노/마이크로/매크로), 지인 추천 바이럴 설계 방법."
                                    )},
                                ],
                                temperature=0.8, max_tokens=2000,
                            )
                            st.session_state.magic_prompt = resp.choices[0].message.content.strip()
                            st.session_state.extra_notes  = st.session_state.magic_prompt
                            st.success("✅ 상세 마케팅 전략 자동 추천 완료!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"AI 추천 실패: {e}")

            extra_notes = st.text_area(
                "📄 AI 추천 마케팅 메모 (자동 채워짐 / 직접 수정 가능)",
                placeholder="위 버튼으로 자동 추천받거나 직접 입력하세요.",
                height=130, key="extra_notes",
            )
            st.markdown('</div>', unsafe_allow_html=True)

        with inp_b:
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown("#### 🖼️ 제품 이미지 업로드")
            uploaded_file = st.file_uploader(
                "JPG/PNG/WEBP 업로드 시 AI가 배경을 자동 분석·합성합니다.",
                type=["jpg","jpeg","png","webp"], key="product_image",
            )
            if uploaded_file:
                st.image(Image.open(uploaded_file), caption="업로드된 원본", use_container_width=True)
                st.success("✅ 이미지 업로드 완료!")

                # 이미지 분석 & 자동 기입
                if st.button("🔍 이미지 AI 분석 → 제품 정보 자동 기입", type="primary", use_container_width=True):
                    with st.spinner("GPT-4o Vision으로 이미지 분석 중 (한국어)..."):
                        uploaded_file.seek(0)
                        img_bytes = uploaded_file.read()
                        nobg_bytes = remove_background(img_bytes)
                        st.session_state.data["raw_img_b64"] = base64.b64encode(img_bytes).decode()
                        st.session_state.data["nobg_img_b64"] = base64.b64encode(nobg_bytes).decode()
                        analysis = vision_analyze_image(img_bytes)
                        st.session_state.data["vision_analysis"] = analysis
                        # ★ Streamlit 제약: 위젯 렌더링 후 동일 key session_state 수정 불가
                        # 해결: 기존 폼 key를 먼저 삭제하고, _pending_autofill에 값 저장
                        # → 다음 rerun 사이클 시작 시 폼 위젯 렌더링 전에 key 세팅
                        for _k in ["in_brand","in_product","in_cat","in_price","in_promo"]:
                            st.session_state.pop(_k, None)
                        st.session_state["_pending_autofill"] = analysis
                        st.rerun()

            else:
                st.info("📌 이미지 미업로드 시 텍스트 기반으로 이미지를 생성합니다.")

            # 분석 결과 표시
            analysis = st.session_state.data.get("vision_analysis", {})
            if analysis:
                with st.expander("🔎 AI 분석 결과 보기", expanded=True):
                    st.markdown(f"**제품명**: {analysis.get('product_name','')}")
                    st.markdown(f"**카테고리**: {analysis.get('category','')}")
                    st.markdown(f"**소재**: {analysis.get('material','')}")
                    st.markdown(f"**색상 팔레트**: {analysis.get('color_palette','')}")
                    st.markdown(f"**타겟**: {analysis.get('target','')}")
                    st.markdown("**핵심 강점**:")
                    for b in analysis.get("key_benefits", []):
                        st.markdown(f"  - {b}")

            st.markdown("---")
            st.markdown("**🎛️ 이미지 합성 모드**")
            synth_mode = st.radio(
                "합성",
                options=["preserve","creative"],
                format_func=lambda x: "✂️ 제품 원본 보존 (누끼 추출 + 배경 합성)" if x=="preserve" else "🎨 AI 창의적 재생성 (형태 변형 가능)",
                index=0, key="synth_mode", label_visibility="collapsed",
            )
            rembg_badge = "✅ rembg 활성" if REMBG_AVAILABLE else "🔧 프롬프트 강화 모드"
            st.caption(f"원본 보존 모드: {rembg_badge}" if synth_mode=="preserve" else "AI가 자유롭게 재해석합니다.")
            st.markdown('</div>', unsafe_allow_html=True)

        # 입력값 보정 (vision_analysis 우선 적용)
        va = st.session_state.data.get("vision_analysis", {})
        final_brand   = input_brand.strip()   or va.get("brand_name","") or m1.get("brand_name","")
        final_product = input_product.strip() or va.get("product_name","") or m1.get("product_name","")
        final_cat     = input_category.strip()or va.get("category","")    or m1.get("category","")
        final_price   = input_price.strip()   or va.get("price_hint","")  or m1.get("price","")
        final_tone    = tone_style.strip()    or va.get("tone","")

        can_go = bool(final_product) and bool(final_brand)
        if not can_go:
            st.caption("⚠️ 브랜드명과 제품명을 입력해야 생성 버튼이 활성화됩니다.")

        st.markdown("")
        if st.button("🚀 전체 마케팅 콘텐츠 자동 생성 (이미지 + 카피 + 영상 프롬프트)", type="primary",
                      use_container_width=True, disabled=not can_go):
            # 이미지 b64 준비
            img_b64 = None
            if "raw_img_b64" in st.session_state.data:
                img_b64 = st.session_state.data["raw_img_b64"]
            elif uploaded_file:
                uploaded_file.seek(0)
                img_b64 = pil_to_b64(Image.open(uploaded_file))

            try:
                with st.status("🔍 제품 브리프 분석 중...", expanded=False) as s:
                    brief = build_brief_cached(
                        final_brand, final_product, final_cat,
                        final_price, input_promo.strip(), extra_notes, final_tone, persona.strip()
                    )
                    st.session_state.ss_brief = brief
                    s.update(label="✅ 브리프 분석 완료", state="complete")

                n_th = 2 if fast_mode else 4
                with st.status(f"🎨 AI 마케팅 이미지 {n_th}가지 테마 생성 중 (약 30~90초)...", expanded=False) as s:
                    themed_raw = generate_images_cached(
                        img_b64, brief.model_dump_json(), image_size, synth_mode, fast_mode
                    )
                    st.session_state.ss_themed_raw  = themed_raw
                    st.session_state.ss_use_rembg   = REMBG_AVAILABLE and synth_mode=="preserve"
                    st.session_state.ss_uploaded    = bool(img_b64)
                    ok = sum(1 for _,_,b in themed_raw if b)
                    s.update(label=f"✅ 이미지 생성 완료 ({ok}/{n_th}장)", state="complete" if ok else "error")

                with st.status("✍️ 플랫폼 맞춤 콘텐츠 생성 중...", expanded=False) as s:
                    pack = build_creative_pack_cached(brief.model_dump_json(), sel_k, final_tone)
                    st.session_state.ss_pack = pack
                    s.update(label="✅ 콘텐츠 생성 완료", state="complete")

                with st.status("🤖 DeepSeek 문구/해시태그 추천 중...", expanded=False) as s:
                    ds = generate_deepseek_copies(brief, sel_k)
                    st.session_state.ss_deepseek = ds
                    s.update(label="✅ DeepSeek 추천 완료", state="complete")

                st.session_state.is_generated = True
                st.session_state.overlay_state = {}
                st.session_state.wanghong_proposals = {}
                st.success("🎉 모든 마케팅 콘텐츠 생성 완료!")
                st.session_state.step = 3
                st.rerun()

            except Exception as e:
                st.error(f"❌ 오류 발생: {e}")
                with st.expander("상세 오류"):
                    st.code(traceback.format_exc())

        st.markdown("---")
        if st.button("← 플랫폼 다시 선택"):
            # 폼 필드 초기화 (다른 제품으로 다시 시작할 때 이전 값 제거)
            for k in ["in_brand","in_product","in_cat","in_price","in_promo",
                      "_autofill_persona","_autofill_tone","_pending_autofill"]:
                st.session_state.pop(k, None)
            st.session_state.step = 1; st.rerun()

    # ══════════════════════════════════════════════════════════════
    # STEP 3: 결과 대시보드
    # ══════════════════════════════════════════════════════════════
    elif step == 3:
        if not st.session_state.get("is_generated"):
            st.warning("먼저 STEP 2에서 콘텐츠를 생성해주세요.")
            if st.button("← STEP 2로"):
                st.session_state.step = 2; st.rerun()
            st.stop()

        sel_k        = st.session_state.data.get("platform","xiaohongshu")
        p_info       = PLATFORMS[sel_k]
        brief        = st.session_state.ss_brief
        pack         = st.session_state.ss_pack
        themed_raw   = st.session_state.ss_themed_raw
        ds           = st.session_state.ss_deepseek
        use_rembg    = st.session_state.ss_use_rembg
        themed_images = [(n,d, b64_to_pil(b) if b else None) for n,d,b in themed_raw]

        st.markdown(f"## 🎉 {p_info['name_cn']} ({p_info['name_kr']}) 마케팅 패키지 완성")
        st.markdown(
            f'<span class="badge b-red">📱 {p_info["name_cn"]}</span> '
            f'<span class="badge b-blue">🛍️ {brief.product_name}</span> '
            f'<span class="badge b-green">✅ 생성 완료</span>',
            unsafe_allow_html=True
        )
        st.markdown("")

        # 컴플라이언스 결과
        violations = getattr(pack, "_compliance_violations", [])
        if violations:
            exp = [BANNED_EXPLANATION.get(v,v) for v in violations]
            st.error("🚨 **[중국 광고법 위반 주의]** 계정 차단 위험 금지어 발견!\n\n" + "\n".join(f"- {e}" for e in exp))
        else:
            st.success("✅ 중국 광고법 필수 검수 통과 (금지어 없음)")

        # 자동 분석 브리프 요약
        with st.expander("📋 자동 분석된 제품 브리프", expanded=False):
            b1, b2 = st.columns(2)
            with b1:
                st.markdown(f"**브랜드명**: {brief.brand_name}")
                st.markdown(f"**제품명**: {brief.product_name}")
                st.markdown(f"**카테고리**: {brief.category}")
                st.markdown(f"**타겟 고객**: {brief.target}")
                st.markdown(f"**가격**: {brief.price or '미입력'}")
                st.markdown(f"**프로모션**: {brief.promo or '미입력'}")
            with b2:
                st.markdown(f"**마케팅 톤**: {brief.tone}")
                st.markdown(f"**CTA**: {brief.landing_action}")
                st.markdown("**핵심 소구점**:")
                for b_ in brief.key_benefits:
                    st.markdown(f"  - {b_}")
                st.code(" / ".join(brief.banned_claims[:5]), language=None)

        st.divider()

        # ── 탭 레이아웃 ──────────────────────────────────────────
        tab1, tab2, tab3, tab4 = st.tabs([
            "🖼️ 이미지 & 문구 합성",
            "📝 마케팅 카피라이팅",
            "🎬 숏폼 영상 프롬프트",
            "💌 왕홍 협업 제안서",
        ])

        # ── TAB 1: 이미지 갤러리 + DeepSeek 문구/해시태그 합성 ──
        with tab1:
            st.markdown("#### 🎨 AI 이미지 갤러리 (소형 미리보기 — 2×2)")
            mode_label = ("✂️ rembg 누끼+배경합성" if use_rembg
                          else ("🔧 프롬프트 강화 원본 보존" if st.session_state.ss_uploaded else "📝 텍스트 프롬프트"))
            st.caption(f"합성 방식: **{mode_label}** | 브리프 오버레이: {'ON' if overlay_on else 'OFF'}")

            overlays = ds.get("overlays", [{"zh":"热卖中","kr":"절찬 판매중"}]*4)

            # 2×2 소형 그리드
            for row_i in range(0, len(themed_images), 2):
                row_items = themed_images[row_i:row_i+2]
                cols = st.columns(2, gap="medium")
                for col, (t_name, t_desc, t_img) in zip(cols, row_items):
                    idx = themed_images.index((t_name, t_desc, t_img))
                    with col:
                        with st.container(border=True):
                            st.markdown(f"**{t_name}**")
                            st.caption(t_desc)
                            if t_img:
                                # 현재 표시할 이미지 (합성 여부에 따라)
                                overlay_key = f"ov_{idx}"
                                disp_b64 = st.session_state.overlay_state.get(overlay_key)
                                if disp_b64:
                                    disp_img = b64_to_pil(disp_b64)
                                else:
                                    disp_img = overlay_brief_on_image(t_img, brief, pack) if overlay_on else t_img
                                # 소형 표시
                                st.image(disp_img, use_container_width=True)

                                # ── 문구 옵션 선택 UI ──────────────────────────
                                st.markdown(
                                    "<div style='font-size:12px;font-weight:700;color:#374151;"
                                    "margin:8px 0 4px;'>📝 합성할 문구 선택 또는 직접 입력</div>",
                                    unsafe_allow_html=True
                                )

                                # 옵션 목록: DeepSeek 4개 추천 + 직접입력
                                option_labels = [f"✏️ 직접 입력"] + [
                                    f"옵션 {i+1}: {ov['zh']} (🇰🇷 {ov['kr']})"
                                    for i, ov in enumerate(overlays)
                                ]
                                sel_opt = st.selectbox(
                                    "문구 선택", option_labels,
                                    key=f"sel_opt_{idx}", label_visibility="collapsed"
                                )

                                if sel_opt.startswith("✏️"):
                                    # 직접 입력
                                    custom_text = st.text_input(
                                        "직접 입력 (중국어)",
                                        placeholder="예: 韩国必买单品",
                                        key=f"custom_txt_{idx}",
                                        label_visibility="collapsed"
                                    )
                                    text_to_synth = custom_text.strip()
                                else:
                                    # 선택된 옵션의 zh 텍스트 추출
                                    opt_idx = option_labels.index(sel_opt) - 1
                                    ov_sel = overlays[opt_idx] if opt_idx < len(overlays) else overlays[0]
                                    text_to_synth = ov_sel["zh"]
                                    # 선택된 문구 미리보기
                                    st.markdown(
                                        f"<div style='background:#f0fdf4;border-radius:8px;"
                                        f"padding:7px 10px;margin:4px 0;'>"
                                        f"<span style='font-size:14px;font-weight:700;color:#166534;'>🇨🇳 {ov_sel['zh']}</span>"
                                        f"<span style='font-size:12px;color:#6b7280;margin-left:8px;'>🇰🇷 {ov_sel['kr']}</span>"
                                        f"</div>",
                                        unsafe_allow_html=True
                                    )

                                c_a, c_b, c_c = st.columns(3)
                                with c_a:
                                    if st.button("✨ 문구 합성", key=f"synth_{idx}", use_container_width=True,
                                                  disabled=not text_to_synth):
                                        raw_b64 = themed_raw[idx][2]
                                        if raw_b64:
                                            st.session_state.overlay_state[overlay_key] = overlay_text_on_image(raw_b64, text_to_synth)
                                            st.rerun()
                                with c_b:
                                    if st.button("🔄 원본", key=f"reset_{idx}", use_container_width=True):
                                        st.session_state.overlay_state.pop(overlay_key, None)
                                        st.rerun()
                                with c_c:
                                    # 다운로드
                                    buf = BytesIO(); disp_img.save(buf, format="PNG")
                                    safe = re.sub(r"[^\w가-힣]","",t_name).strip()
                                    st.download_button("⬇️", data=buf.getvalue(),
                                                        file_name=f"{brief.product_name}_{safe}.png",
                                                        mime="image/png", key=f"dl_{idx}", use_container_width=True)
                            else:
                                st.warning("이미지 생성 실패")

            st.divider()

            # DeepSeek 해시태그 추천
            st.markdown("#### #️⃣ DeepSeek 해시태그 추천")
            h_cn = ds.get("hashtags_cn", [])
            h_kr = ds.get("hashtags_kr", [])
            hc1, hc2 = st.columns(2)
            with hc1:
                st.markdown("🇨🇳 **중국어 해시태그**")
                hash_cn_txt = " ".join(h_cn)
                st.text_area("CN 해시태그", value=hash_cn_txt, height=80, label_visibility="collapsed")
                copy_btn(hash_cn_txt, "cp_hash_cn", "📋 해시태그 복사", True)
            with hc2:
                st.markdown("🇰🇷 **한국어 번역** (참고용)")
                hash_kr_txt = " ".join(h_kr)
                st.text_area("KR 해시태그", value=hash_kr_txt, height=80, label_visibility="collapsed")

            st.info("💡 이미지를 Canva에서 추가 편집하거나, STEP 2에서 이미지 규격을 변경 후 재생성할 수 있습니다.")

        # ── TAB 2: 마케팅 카피라이팅 ───────────────────────────
        with tab2:
            st.markdown(f"#### 📝 {p_info['name_cn']} 맞춤 마케팅 문구")

            current_body_cn = st.session_state.retranslated.get(sel_k, pack.body_cn)

            mc1, mc2 = st.columns(2, gap="large")
            with mc1:
                st.markdown("##### 🇨🇳 중국어 원문")
                with st.container(border=True):
                    st.markdown(f"**제목**: {pack.title_cn}")
                    copy_btn(pack.title_cn, "cp_t_cn")
                    st.markdown(f"**핵심 캡션**: {pack.hook_cn}")
                    st.markdown(f"**CTA**: {pack.cta_cn}")
                    st.markdown(f"**썸네일 문구**: {pack.thumbnail_text_cn}")
                st.markdown("**본문 (중국어)**")
                st.text_area("cn_body", value=current_body_cn, height=200, label_visibility="collapsed")
                copy_btn(current_body_cn, "cp_body_cn", "📋 본문 복사", True)
                st.markdown("**해시태그**")
                hash_str = " ".join(pack.hashtags)
                st.text_area("cn_hash", value=hash_str, height=70, label_visibility="collapsed")
                copy_btn(hash_str, "cp_pack_hash", "📋 해시태그 복사", True)

                # 전체 복사
                full_cn = f"{pack.title_cn}\n\n{pack.hook_cn}\n\n{current_body_cn}\n\n{pack.cta_cn}\n\n{hash_str}"
                copy_btn(full_cn, "cp_full_cn", "📋 전체 복사", True)

            with mc2:
                st.markdown("##### 🇰🇷 한국어 번역 (수정 후 중국어 재번역 가능)")
                with st.container(border=True):
                    st.markdown(f"**제목**: {pack.title_kr}")
                    st.markdown(f"**핵심 캡션**: {pack.hook_kr}")
                    st.markdown(f"**CTA**: {pack.cta_kr}")
                    st.markdown(f"**썸네일 문구**: {pack.thumbnail_text_kr}")
                st.markdown("**본문 (한국어) — 수정 후 재번역 가능**")
                kr_key = f"kr_edit_{sel_k}"
                if kr_key not in st.session_state:
                    st.session_state[kr_key] = pack.body_kr
                edited_kr = st.text_area("kr_body", height=200, key=kr_key, label_visibility="collapsed")
                if st.button("🔄 수정한 한국어로 중국어 재번역", key=f"retrans_{sel_k}", use_container_width=True, type="secondary"):
                    with st.spinner("중국어 재번역 중..."):
                        try:
                            new_cn = retranslate_body_cn(edited_kr, brief, sel_k)
                            st.session_state.retranslated[sel_k] = new_cn
                            st.success("✅ 재번역 완료!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"재번역 실패: {e}")

            st.divider()

            # 숏폼 자막
            st.markdown("#### 🧷 숏폼 영상 자막")
            sc1, sc2 = st.columns(2)
            with sc1:
                st.markdown("🇨🇳 **중국어 자막**")
                subs_cn = "\n".join(f"{i}. {s}" for i,s in enumerate(pack.subtitles_cn, 1))
                st.text_area("sub_cn", value=subs_cn, height=160, label_visibility="collapsed")
                copy_btn(subs_cn, "cp_subs_cn", "📋 자막 복사", True)
            with sc2:
                st.markdown("🇰🇷 **한국어 번역**")
                subs_kr = "\n".join(f"{i}. {s}" for i,s in enumerate(pack.subtitles_kr, 1))
                st.text_area("sub_kr", value=subs_kr, height=160, label_visibility="collapsed")

        # ── TAB 3: 숏폼 영상 프롬프트 ─────────────────────────
        with tab3:
            st.markdown(f"#### 🎬 숏폼 영상 스토리보드 (씬별 기획안)")
            st.caption(f"권장 영상 길이: {p_info['video_duration']} | 플랫폼: {p_info['name_cn']}")

            if pack.storyboard:
                # 전체 씬 복사 (상단 배치)
                all_scenes_text = ""
                for sc in pack.storyboard:
                    all_scenes_text += (
                        f"[{sc.get('scene','')}] ⏱ {sc.get('duration','')}\n"
                        f"📷 화면: {sc.get('visual','')}\n"
                        f"🇨🇳 자막: {sc.get('caption_cn','')}\n"
                        f"🇰🇷 번역: {sc.get('caption_kr','')}\n\n"
                    )
                copy_btn(all_scenes_text.strip(), "cp_all_scenes", "📋 전체 씬 복사", True)
                st.markdown("")

                for i, sc in enumerate(pack.storyboard, 1):
                    with st.container(border=True):
                        sc1, sc2 = st.columns([1, 3])
                        with sc1:
                            st.markdown(f"**씬 {i}**")
                            st.caption(f"⏱ {sc.get('duration','')}")
                        with sc2:
                            st.markdown(f"📷 **화면 구성**: {sc.get('visual','')}")
                            st.markdown(f"🇨🇳 **자막 (CN)**: `{sc.get('caption_cn','')}`")
                            st.markdown(f"🇰🇷 **번역 (KR)**: {sc.get('caption_kr','')}")
                            # 씬별 프롬프트 복사
                            scene_txt = f"[{sc.get('scene',f'씬 {i}')}]\n화면: {sc.get('visual','')}\n자막(CN): {sc.get('caption_cn','')}\n자막(KR): {sc.get('caption_kr','')}"
                            copy_btn(scene_txt, f"cp_scene_{i}", "📋 이 씬 복사")
            else:
                st.info("스토리보드 생성에 실패했습니다. STEP 2로 돌아가 재생성해주세요.")

            st.divider()
            st.markdown("#### 🚀 AI 영상 제작 툴 바로 가기 (6종)")
            st.caption("아래 도구들에 위 프롬프트를 입력해 숏폼 영상을 제작하세요.")

            # 6개 툴을 3+3으로 배치
            tool_row1 = VIDEO_TOOLS[:3]
            tool_row2 = VIDEO_TOOLS[3:]
            for tool_row in [tool_row1, tool_row2]:
                row_cols = st.columns(3, gap="medium")
                for col, tool in zip(row_cols, tool_row):
                    with col:
                        with st.container(border=True):
                            st.markdown(f"**{tool['icon']} {tool['name']}**")
                            st.caption(f"💰 {tool['price']}")
                            st.markdown(f"<div style='font-size:12px;color:#4b5563;min-height:48px;'>{tool['desc']}</div>", unsafe_allow_html=True)
                            st.markdown(
                                f"<a href='{tool['url']}' target='_blank' style='display:inline-block;margin-top:8px;"
                                f"padding:7px 14px;background:#e53935;color:#fff;border-radius:8px;"
                                f"text-decoration:none;font-size:12px;font-weight:700;'>🔗 바로 가기</a>",
                                unsafe_allow_html=True
                            )

        # ── TAB 4: 왕홍 협업 제안서 ───────────────────────────
        with tab4:
            st.markdown("#### 💌 왕홍(KOL/KOC) 협업 제안서 3종 자동 생성")
            st.caption("DM 짧은버전 / DM 긴버전 / 이메일 — 한국어 + 중국어 쌍으로 생성됩니다.")

            # 협업 조건 입력
            with st.container(border=True):
                st.markdown("**👤 왕홍 정보 (선택 — 입력 시 맞춤 제안서 생성)**")
                wh_col1, wh_col2, wh_col3 = st.columns(3)
                with wh_col1:
                    wanghong_name    = st.text_input("왕홍 이름/닉네임 *", placeholder="예: 李佳琦, 小红书美妆博主", key="wh_name")
                with wh_col2:
                    wanghong_tier    = st.selectbox("왕홍 티어", ["나노 (팔로워 1만 미만)", "마이크로 (1만~10만)", "매크로 (10만~100만)", "메가 (100만+)"], key="wh_tier")
                with wh_col3:
                    wanghong_style   = st.text_input("왕홍 콘텐츠 스타일", placeholder="예: 뷰티 리뷰, 일상 브이로그", key="wh_style")

                st.markdown("**📋 협업 조건 설정**")
                c1, c2, c3, c4 = st.columns(4)
                with c1: collab_count    = st.text_input("콘텐츠 수", value="1~2개", key="c_count")
                with c2: collab_format   = st.text_input("형식", value="图文+Reel", key="c_format")
                with c3: collab_deadline = st.text_input("납기", value="협의 후 결정", key="c_deadline")
                with c4: collab_benefit  = st.text_input("제공 혜택", value="제품 무료 제공 + 커미션", key="c_benefit")

            if st.button("✉️ 3종 협업 제안서 자동 생성", use_container_width=True, type="primary"):
                with st.spinner("협업 제안서 작성 중..."):
                    try:
                        proposals = generate_wanghong_proposal(
                            brief, sel_k, collab_count, collab_format, collab_deadline, collab_benefit,
                            wanghong_name=wanghong_name.strip(), wanghong_tier=wanghong_tier, wanghong_style=wanghong_style.strip()
                        )
                        st.session_state.wanghong_proposals = proposals
                        st.success(f"✅ {'[' + wanghong_name + '] ' if wanghong_name else ''}제안서 생성 완료!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"제안서 생성 실패: {e}")

            proposals = st.session_state.get("wanghong_proposals", {})
            if proposals:
                pt1, pt2, pt3 = st.tabs(["📩 DM 짧은버전 (100~150자)", "📨 DM 긴버전 (200~300자)", "📧 이메일 버전"])
                with pt1:
                    pc1, pc2 = st.columns(2)
                    with pc1:
                        st.markdown("🇰🇷 한국어")
                        st.text_area("ds_kr", value=proposals.get("dm_short_kr",""), height=120, label_visibility="collapsed")
                        copy_btn(proposals.get("dm_short_kr",""), "cp_ds_kr", "📋 복사")
                    with pc2:
                        st.markdown("🇨🇳 중국어")
                        st.text_area("ds_cn", value=proposals.get("dm_short_cn",""), height=120, label_visibility="collapsed")
                        copy_btn(proposals.get("dm_short_cn",""), "cp_ds_cn", "📋 복사")
                with pt2:
                    pc1, pc2 = st.columns(2)
                    with pc1:
                        st.markdown("🇰🇷 한국어")
                        st.text_area("dl_kr", value=proposals.get("dm_long_kr",""), height=160, label_visibility="collapsed")
                        copy_btn(proposals.get("dm_long_kr",""), "cp_dl_kr", "📋 복사")
                    with pc2:
                        st.markdown("🇨🇳 중국어")
                        st.text_area("dl_cn", value=proposals.get("dm_long_cn",""), height=160, label_visibility="collapsed")
                        copy_btn(proposals.get("dm_long_cn",""), "cp_dl_cn", "📋 복사")
                with pt3:
                    pc1, pc2 = st.columns(2)
                    with pc1:
                        st.markdown("🇰🇷 한국어")
                        st.markdown(f"**제목**: {proposals.get('email_subject_kr','')}")
                        st.text_area("eb_kr", value=proposals.get("email_body_kr",""), height=240, label_visibility="collapsed")
                        copy_btn(proposals.get("email_body_kr",""), "cp_eb_kr", "📋 복사")
                    with pc2:
                        st.markdown("🇨🇳 중국어")
                        st.markdown(f"**主题**: {proposals.get('email_subject_cn','')}")
                        st.text_area("eb_cn", value=proposals.get("email_body_cn",""), height=240, label_visibility="collapsed")
                        copy_btn(proposals.get("email_body_cn",""), "cp_eb_cn", "📋 복사")
            else:
                st.info("위 버튼을 눌러 협업 제안서를 생성하세요.")

        st.divider()
        nc1, nc2 = st.columns(2)
        with nc1:
            if st.button("← STEP 2로 돌아가기"):
                st.session_state.step = 2; st.rerun()
        with nc2:
            if st.button("📊 STEP 4 — 다운로드 →", type="primary"):
                st.session_state.step = 4; st.rerun()

        # ── 빠른 엑셀 다운로드 (STEP 3 하단 바로 가기) ──────────
        st.divider()
        with st.expander("📊 종합 기획안 빠른 다운로드 (STEP 4 이동 없이)", expanded=False):
            st.caption("왕홍 제안서를 먼저 생성하면 엑셀에 자동 포함됩니다.")
            proposals_now = st.session_state.get("wanghong_proposals", {})
            if not proposals_now:
                st.info("💡 TAB4(왕홍 협업 제안서)에서 제안서를 생성하면 엑셀에 포함됩니다.")
            dl_c1, dl_c2 = st.columns(2)
            with dl_c1:
                excel_bytes_quick = generate_excel_report(brief, pack, proposals_now, themed_raw, sel_k)
                st.download_button(
                    "📊 종합 광고 기획안 다운로드 (Excel)",
                    data=excel_bytes_quick,
                    file_name=f"왕홍마케팅_기획안_{brief.product_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, type="primary",
                )
            with dl_c2:
                export_quick = {"brief": brief.model_dump(), "pack": pack.model_dump(), "proposals": proposals_now}
                st.download_button(
                    "⬇️ 원본 데이터 JSON",
                    data=json.dumps(export_quick, ensure_ascii=False, indent=2).encode("utf-8"),
                    file_name=f"wanghong_{brief.product_name}_output.json",
                    mime="application/json", use_container_width=True,
                )

    # ══════════════════════════════════════════════════════════════
    # STEP 4: 다운로드
    # ══════════════════════════════════════════════════════════════
    elif step == 4:
        if not st.session_state.get("is_generated"):
            st.warning("먼저 콘텐츠를 생성해주세요.")
            if st.button("← 처음으로"):
                st.session_state.step = 1; st.rerun()
            st.stop()

        sel_k    = st.session_state.data.get("platform","xiaohongshu")
        brief    = st.session_state.ss_brief
        pack     = st.session_state.ss_pack
        themed_raw = st.session_state.ss_themed_raw

        st.markdown("## 📊 STEP 4 — 종합 기획안 다운로드")
        st.caption("엑셀 파일을 열어 전체 내용 확인·수정 후 실무에 바로 활용하세요.")

        proposals = st.session_state.get("wanghong_proposals", {})
        if not proposals:
            st.info("💡 왕홍 협업 제안서를 먼저 생성하면 엑셀에 포함됩니다. (STEP 3 → TAB4)")

        dc1, dc2 = st.columns(2, gap="large")
        with dc1:
            excel_bytes = generate_excel_report(brief, pack, proposals, themed_raw, sel_k)
            st.download_button(
                "📊 종합 광고 기획안 다운로드 (Excel)",
                data=excel_bytes,
                file_name=f"왕홍마케팅_기획안_{brief.product_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, type="primary",
            )
        with dc2:
            export = {"brief": brief.model_dump(), "pack": pack.model_dump(), "proposals": proposals}
            st.download_button(
                "⬇️ 원본 데이터 JSON 다운로드",
                data=json.dumps(export, ensure_ascii=False, indent=2).encode("utf-8"),
                file_name=f"wanghong_{brief.product_name}_output.json",
                mime="application/json", use_container_width=True,
            )

        st.markdown("---")
        fc1, fc2 = st.columns(2)
        with fc1:
            if st.button("← STEP 3 결과로 돌아가기"):
                st.session_state.step = 3; st.rerun()
        with fc2:
            if st.button("🏠 처음부터 다시 시작", type="secondary"):
                for k in list(st.session_state.keys()):
                    del st.session_state[k]
                st.rerun()

if __name__ == "__main__":
    main()